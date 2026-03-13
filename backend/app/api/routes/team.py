import secrets
from datetime import datetime, timezone
from io import BytesIO
import io
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Path, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from openpyxl import Workbook, load_workbook

from app.api.deps import get_current_user, require_page_permission
from app.core.permissions import BASE_DEFAULT_PERMISSIONS_BY_ROLE, PERMISSIONS, PERMISSION_KEYS
from app.core.security import hash_password
from app.db.session import get_db
from app.models.project import Project
from app.models.shop_point import ShopPoint
from app.models.user import User
from app.models.user_point_access import UserPointAccess
from app.models.user_project_access import UserProjectAccess
from app.models.user_role import UserRole
from app.models.user_group import UserGroup
from app.models.user_group_member import UserGroupMember
from app.models.user_company_access import UserCompanyAccess
from app.models.user_company_distribution import UserCompanyDistribution
from app.models.owner_company_access import OwnerCompanyAccess
from app.models.company import Company
from app.models.company_settings import CompanySettings
from app.models.user_company_reports import UserCompanyReports
from app.schemas.access import (
  BulkAssignRequest,
  BulkAssignResponse,
  UserPointAccessReplaceRequest,
  UserPointAccessResponse,
  UserProjectAccessReplaceRequest,
  UserProjectAccessResponse,
)
from app.schemas.projects import ShopPointsResponse
from app.api.routes.projects import to_shop_point_item  # type: ignore
from app.schemas.company import CompanyPublic
from app.schemas.team import (
  TeamCreateUserRequest,
  TeamCreateUserResponse,
  TeamResetPasswordResponse,
  TeamUpdateUserRequest,
  TeamUserDetailsResponse,
  TeamUserItem,
  PermissionItem,
  RoleDefaultPermissionsResponse,
  RoleDefaultPermissionsUpdateRequest,
  TeamUsersImportError,
  TeamUsersImportResponse,
  TeamUsersResponse,
)
from app.schemas.groups import (
  AddUserGroupMemberByEmailRequest,
  CreateUserGroupRequest,
  ReplaceUserGroupMembersRequest,
  ReplaceUserGroupsForUserRequest,
  UpdateUserGroupRequest,
  UserGroupItem,
  UserGroupMembersImportError,
  UserGroupMembersImportResponse,
  UserGroupMembersResponse,
  UserGroupsForUserResponse,
  UserGroupsResponse,
)
from app.schemas.distribution import (
  CompanyFilterValuesResponse,
  DistributionCompanyItem,
  ReplaceUserCompaniesAccessRequest,
  ReplaceUserCompanyDistributionRequest,
  UserCompaniesAccessResponse,
  UserCompanyDistributionResponse,
)
from app.schemas.reports_access import ReplaceUserCompanyReportsRequest, UserCompanyReportsResponse
from app.schemas.user_clone import CloneUserRequest, CloneUserResponse


router = APIRouter(dependencies=[Depends(require_page_permission('page.team.view'))])


def to_team_user_item (user: User) -> TeamUserItem:
  company = user.company
  if company is None:
    company_public = CompanyPublic(id=user.company_id, name='—')
  else:
    company_public = CompanyPublic(id=company.id, name=company.name)

  return TeamUserItem(
    id=user.id,
    publicId=int(user.public_id),
    firstName=user.first_name,
    lastName=user.last_name,
    email=user.email,
    phone=user.phone,
    company=company_public,
    role=user.role,
    profileCompany=user.profile_company,
    uiLanguage=user.ui_language,
    isActive=bool(user.is_active),
    permissions=list(user.permissions or []),
    note=user.note,
    createdAt=user.created_at,
    lastLoginAt=user.last_login_at,
  )


ROLE_PERMISSIONS_THEME_KEY = 'roleDefaultPermissions'


def _normalize_permissions (values: list[str] | None, *, role: UserRole, defaults_by_role: dict[UserRole, list[str]]) -> list[str]:
  if not values:
    return list(defaults_by_role.get(role, []))

  out: list[str] = []
  seen: set[str] = set()
  for raw in values:
    k = str(raw or '').strip()
    if not k:
      continue
    if k not in PERMISSION_KEYS:
      raise HTTPException(status_code=400, detail=f'Некорректный permission: {k}')
    if k not in seen:
      seen.add(k)
      out.append(k)
  return out


def _ensure_admin_or_owner (user: User) -> None:
  if (user.platform_role or 'user') == 'owner':
    return
  if user.role == UserRole.admin:
    return
  raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail='Недостаточно прав')


ROLE_LEVEL: dict[UserRole, int] = {
  UserRole.admin: 1,
  UserRole.manager: 2,
  UserRole.controller: 3,
  UserRole.coordinator: 4,
  UserRole.client: 5,
}


def _now_utc () -> datetime:
  return datetime.now(timezone.utc)


def _ensure_can_manage_user (actor: User, *, target_role: UserRole) -> None:
  if (actor.platform_role or 'user') == 'owner':
    return
  if actor.role == UserRole.admin:
    return
  if actor.role == UserRole.manager:
    if target_role == UserRole.admin:
      raise HTTPException(status_code=403, detail='Менеджер не может управлять администратором')
    return
  raise HTTPException(status_code=403, detail='Недостаточно прав')


def _ensure_admin (user: User) -> None:
  if user.role != UserRole.admin:
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail='Недостаточно прав')

def _ensure_admin_or_manager_or_owner (user: User) -> None:
  if (user.platform_role or 'user') == 'owner':
    return
  if user.role not in (UserRole.admin, UserRole.manager):
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail='Недостаточно прав')


def _has_perm (user: User, key: str) -> bool:
  return key in set(user.permissions or [])


def _ensure_perm (user: User, key: str) -> None:
  if (user.platform_role or 'user') == 'owner':
    return
  if user.role in (UserRole.admin, UserRole.manager) and _has_perm(user, key):
    return
  raise HTTPException(status_code=403, detail='Недостаточно прав')


def _company_id (current_user: User) -> UUID:
  return getattr(current_user, 'active_company_id', current_user.company_id)


def _merge_role_defaults_from_theme (theme: dict | None) -> dict[UserRole, list[str]]:
  out: dict[UserRole, list[str]] = {role: list(items) for role, items in BASE_DEFAULT_PERMISSIONS_BY_ROLE.items()}
  if not isinstance(theme, dict):
    return out
  raw = theme.get(ROLE_PERMISSIONS_THEME_KEY)
  if not isinstance(raw, dict):
    return out

  for role in UserRole:
    role_values = raw.get(role.value)
    if not isinstance(role_values, list):
      continue
    out[role] = _normalize_permissions(role_values, role=role, defaults_by_role=out)
  return out


def _serialize_role_defaults (by_role: dict[UserRole, list[str]]) -> dict[str, list[str]]:
  out: dict[str, list[str]] = {}
  for role in UserRole:
    out[role.value] = list(by_role.get(role, []))
  return out


async def _get_company_role_defaults (db: AsyncSession, *, company_id: UUID) -> tuple[CompanySettings | None, dict[UserRole, list[str]]]:
  res = await db.execute(select(CompanySettings).where(CompanySettings.company_id == company_id))
  settings = res.scalar_one_or_none()
  theme = (settings.theme or {}) if settings is not None else {}
  by_role = _merge_role_defaults_from_theme(theme)
  return settings, by_role


def _normalize_region_codes (codes: list[str]) -> list[str]:
  out: list[str] = []
  seen: set[str] = set()
  for c in codes:
    v = str(c or '').strip()
    if not v:
      continue
    if v == '*':
      return ['*']
    if v.isdigit() and len(v) == 1:
      v = v.zfill(2)
    if v not in seen:
      seen.add(v)
      out.append(v)
  return out


def _normalize_access_role (value: str) -> str:
  v = (value or '').strip().lower()
  if v not in ('controller', 'coordinator'):
    raise HTTPException(status_code=400, detail='Некорректная роль доступа')
  return v


def _xlsx_bytes_from_rows (headers: list[str], rows: list[list[object | None]]) -> bytes:
  wb = Workbook(write_only=True)
  ws = wb.create_sheet(title='Sheet1')
  ws.append(headers)
  for row in rows:
    ws.append([('' if value is None else value) for value in row])

  buf = io.BytesIO()
  wb.save(buf)
  return buf.getvalue()


async def _ensure_project_manage_access (
  db: AsyncSession,
  *,
  actor: User,
  project_id: UUID,
) -> None:
  company_id = _company_id(actor)
  res = await db.execute(select(Project).where(Project.id == project_id, Project.company_id == company_id))
  project = res.scalar_one_or_none()
  if project is None:
    raise HTTPException(status_code=404, detail='Проект не найден')

  if (actor.platform_role or 'user') == 'owner':
    return
  if actor.role == UserRole.admin:
    return
  if actor.role == UserRole.manager and project.manager_user_id == actor.id:
    return
  raise HTTPException(status_code=403, detail='Недостаточно прав')


async def _get_base_ap_project_id (db: AsyncSession, *, company_id: UUID) -> UUID | None:
  res = await db.execute(select(CompanySettings.base_ap_project_id).where(CompanySettings.company_id == company_id))
  return res.scalar_one_or_none()


async def _ensure_target_user_in_company (db: AsyncSession, *, company_id: UUID, user_id: UUID) -> User:
  res = await db.execute(select(User).where(User.id == user_id, User.company_id == company_id))
  user = res.scalar_one_or_none()
  if user is None:
    raise HTTPException(status_code=404, detail='Пользователь не найден')
  return user


@router.get(
  '/users',
  response_model=TeamUsersResponse,
  summary='Список участников',
  description='Возвращает список пользователей в рамках компании текущего пользователя. Поддерживает поиск и фильтр по роли.',
)
async def list_users (
  q: str | None = Query(default=None, description='Поиск по имени, фамилии или почте'),
  user_id: UUID | None = Query(default=None, description='Фильтр по ID пользователя'),
  public_id: int | None = Query(default=None, description='Фильтр по публичному ID пользователя'),
  is_active: bool | None = Query(default=None, description='Фильтр по активности'),
  role: UserRole | None = Query(default=None, description='Фильтр по роли'),
  email: str | None = Query(default=None, description='Фильтр по email'),
  phone: str | None = Query(default=None, description='Фильтр по телефону'),
  profile_company: str | None = Query(default=None, description='Фильтр по компании (профиль)'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> TeamUsersResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = getattr(current_user, 'active_company_id', current_user.company_id)

  stmt = (
    select(User)
    .options(selectinload(User.company))
    .where(User.company_id == company_id)
    .order_by(User.created_at.desc())
  )
  if user_id is not None:
    stmt = stmt.where(User.id == user_id)
  if public_id is not None:
    stmt = stmt.where(User.public_id == public_id)
  if is_active is not None:
    stmt = stmt.where(User.is_active == is_active)
  if role is not None:
    stmt = stmt.where(User.role == role)
  if email:
    stmt = stmt.where(User.email.ilike(f'%{email.strip()}%'))
  if phone:
    stmt = stmt.where(User.phone.ilike(f'%{phone.strip()}%'))
  if profile_company:
    stmt = stmt.where(User.profile_company.ilike(f'%{profile_company.strip()}%'))
  if q:
    like = f'%{q.strip()}%'
    stmt = stmt.where(
      or_(
        User.first_name.ilike(like),
        User.last_name.ilike(like),
        User.email.ilike(like),
      ),
    )

  res = await db.execute(stmt)
  users = res.scalars().all()
  return TeamUsersResponse(items=[to_team_user_item(u) for u in users])


@router.get(
  '/users/export',
  summary='Экспорт участников в Excel (.xlsx)',
)
async def export_users_xlsx (
  q: str | None = Query(default=None, description='Поиск по имени, фамилии или почте'),
  user_id: UUID | None = Query(default=None, description='Фильтр по ID пользователя'),
  public_id: int | None = Query(default=None, description='Фильтр по публичному ID пользователя'),
  is_active: bool | None = Query(default=None, description='Фильтр по активности'),
  role: UserRole | None = Query(default=None, description='Фильтр по роли'),
  email: str | None = Query(default=None, description='Фильтр по email'),
  phone: str | None = Query(default=None, description='Фильтр по телефону'),
  profile_company: str | None = Query(default=None, description='Фильтр по компании (профиль)'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> StreamingResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'data.export')
  company_id = _company_id(current_user)

  stmt = (
    select(User)
    .options(selectinload(User.company))
    .where(User.company_id == company_id)
    .order_by(User.created_at.desc())
  )
  if user_id is not None:
    stmt = stmt.where(User.id == user_id)
  if public_id is not None:
    stmt = stmt.where(User.public_id == public_id)
  if is_active is not None:
    stmt = stmt.where(User.is_active == is_active)
  if role is not None:
    stmt = stmt.where(User.role == role)
  if email:
    stmt = stmt.where(User.email.ilike(f'%{email.strip()}%'))
  if phone:
    stmt = stmt.where(User.phone.ilike(f'%{phone.strip()}%'))
  if profile_company:
    stmt = stmt.where(User.profile_company.ilike(f'%{profile_company.strip()}%'))
  if q:
    like = f'%{q.strip()}%'
    stmt = stmt.where(
      or_(
        User.first_name.ilike(like),
        User.last_name.ilike(like),
        User.email.ilike(like),
      ),
    )

  res = await db.execute(stmt)
  users = res.scalars().all()

  headers = [
    'ID',
    'UUID',
    'Имя',
    'Фамилия',
    'Email',
    'Телефон',
    'Компания',
    'Роль',
    'Компания (профиль)',
    'Язык интерфейса',
    'Активен',
    'Дата создания',
    'Последний вход',
  ]
  rows: list[list[object | None]] = []
  for user in users:
    rows.append(
      [
        int(user.public_id),
        str(user.id),
        user.first_name,
        user.last_name,
        user.email,
        user.phone,
        user.company.name if user.company is not None else None,
        user.role.value if hasattr(user.role, 'value') else str(user.role),
        user.profile_company,
        user.ui_language,
        bool(user.is_active),
        user.created_at.isoformat() if user.created_at else None,
        user.last_login_at.isoformat() if user.last_login_at else None,
      ],
    )

  content = _xlsx_bytes_from_rows(headers, rows)
  filename = f'team-users-{company_id}.xlsx'
  return StreamingResponse(
    io.BytesIO(content),
    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    headers={'Content-Disposition': f'attachment; filename="{filename}"'},
  )


@router.get(
  '/role-default-permissions',
  response_model=RoleDefaultPermissionsResponse,
  summary='Дефолтные права ролей',
)
async def get_role_default_permissions (
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> RoleDefaultPermissionsResponse:
  _ensure_admin_or_owner(current_user)
  _, by_role = await _get_company_role_defaults(db, company_id=_company_id(current_user))
  return RoleDefaultPermissionsResponse(
    byRole=by_role,
    availablePermissions=[PermissionItem(key=key, label=label) for key, label in PERMISSIONS],
  )


@router.put(
  '/role-default-permissions',
  response_model=RoleDefaultPermissionsResponse,
  summary='Обновить дефолтные права ролей',
)
async def update_role_default_permissions (
  payload: RoleDefaultPermissionsUpdateRequest,
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> RoleDefaultPermissionsResponse:
  _ensure_admin_or_owner(current_user)
  company_id = _company_id(current_user)
  settings, _ = await _get_company_role_defaults(db, company_id=company_id)

  normalized: dict[UserRole, list[str]] = {}
  for role in UserRole:
    incoming = payload.byRole.get(role, payload.byRole.get(role.value, []))  # type: ignore[arg-type]
    role_permissions = _normalize_permissions(
      incoming,
      role=role,
      defaults_by_role=BASE_DEFAULT_PERMISSIONS_BY_ROLE,
    )
    if role == UserRole.admin:
      for forced_key in ('page.team.view', 'page.permissions.view', 'users.edit'):
        if forced_key not in role_permissions:
          role_permissions.append(forced_key)
    normalized[role] = role_permissions

  if settings is None:
    settings = CompanySettings(company_id=company_id, theme={})
    db.add(settings)

  next_theme = dict(settings.theme or {})
  next_theme[ROLE_PERMISSIONS_THEME_KEY] = _serialize_role_defaults(normalized)
  settings.theme = next_theme
  await db.commit()

  return RoleDefaultPermissionsResponse(
    byRole=normalized,
    availablePermissions=[PermissionItem(key=key, label=label) for key, label in PERMISSIONS],
  )


@router.post(
  '/users',
  response_model=TeamCreateUserResponse,
  status_code=status.HTTP_201_CREATED,
  summary='Добавить участника',
  description='Создаёт пользователя в компании текущего пользователя. Доступно только администратору.',
)
async def create_user (
  payload: TeamCreateUserRequest,
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> TeamCreateUserResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  _ensure_can_manage_user(current_user, target_role=payload.role)

  existing = await db.execute(select(User).where(User.email == str(payload.email)))
  if existing.scalar_one_or_none() is not None:
    raise HTTPException(status_code=400, detail='Пользователь с такой почтой уже существует')

  _, role_defaults = await _get_company_role_defaults(db, company_id=_company_id(current_user))

  temporary_password = (payload.password or '').strip()
  if not temporary_password:
    temporary_password = secrets.token_urlsafe(9)
    temporary_password = temporary_password[:20]

  user = User(
    first_name=payload.firstName,
    last_name=payload.lastName,
    email=str(payload.email),
    phone=(payload.phone or None),
    role=payload.role,
    note=payload.note,
    profile_company=payload.profileCompany,
    ui_language=(payload.uiLanguage or 'ru'),
    is_active=bool(payload.isActive),
    company_id=getattr(current_user, 'active_company_id', current_user.company_id),
    temporary_password=temporary_password,
    password_hash=hash_password(temporary_password),
    permissions=_normalize_permissions(payload.permissions, role=payload.role, defaults_by_role=role_defaults),
  )
  db.add(user)
  await db.commit()
  res = await db.execute(select(User).options(selectinload(User.company)).where(User.id == user.id))
  user = res.scalar_one()

  return TeamCreateUserResponse(user=to_team_user_item(user), temporaryPassword=temporary_password)


@router.get(
  '/users/{user_id}',
  response_model=TeamUserDetailsResponse,
  summary='Карточка участника',
  description='Возвращает данные пользователя в рамках компании текущего пользователя. Администратор дополнительно получает временный пароль (прототип).',
)
async def get_user (
  user_id: UUID = Path(..., description='ID пользователя'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> TeamUserDetailsResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = getattr(current_user, 'active_company_id', current_user.company_id)
  res = await db.execute(
    select(User)
    .options(selectinload(User.company))
    .where(User.id == user_id, User.company_id == company_id),
  )
  user = res.scalar_one_or_none()
  if user is None:
    raise HTTPException(status_code=404, detail='Пользователь не найден')

  password = user.temporary_password if current_user.role == UserRole.admin else None
  return TeamUserDetailsResponse(user=to_team_user_item(user), password=password)


@router.patch(
  '/users/{user_id}',
  response_model=TeamUserDetailsResponse,
  summary='Обновить участника',
)
async def update_user (
  payload: TeamUpdateUserRequest,
  user_id: UUID = Path(..., description='ID пользователя'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> TeamUserDetailsResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  res = await db.execute(select(User).options(selectinload(User.company)).where(User.id == user_id, User.company_id == company_id))
  user = res.scalar_one_or_none()
  if user is None:
    raise HTTPException(status_code=404, detail='Пользователь не найден')

  _ensure_can_manage_user(current_user, target_role=user.role)
  _, role_defaults = await _get_company_role_defaults(db, company_id=company_id)

  if payload.role is not None:
    _ensure_can_manage_user(current_user, target_role=payload.role)

  if user.id == current_user.id and payload.isActive is False:
    raise HTTPException(status_code=400, detail='Нельзя деактивировать самого себя')

  if payload.email is not None:
    next_email = str(payload.email)
    if next_email != user.email:
      exists = await db.execute(select(User.id).where(User.email == next_email, User.id != user.id))
      if exists.scalar_one_or_none() is not None:
        raise HTTPException(status_code=400, detail='Пользователь с такой почтой уже существует')
      user.email = next_email

  if payload.phone is not None:
    next_phone = payload.phone or None
    if next_phone != user.phone:
      if next_phone:
        exists = await db.execute(select(User.id).where(User.phone == next_phone, User.id != user.id))
        if exists.scalar_one_or_none() is not None:
          raise HTTPException(status_code=400, detail='Пользователь с таким телефоном уже существует')
      user.phone = next_phone

  if payload.firstName is not None:
    user.first_name = payload.firstName
  if payload.lastName is not None:
    user.last_name = payload.lastName
  if payload.role is not None:
    user.role = payload.role
  if payload.note is not None:
    user.note = payload.note
  if payload.profileCompany is not None:
    user.profile_company = payload.profileCompany
  if payload.uiLanguage is not None:
    user.ui_language = payload.uiLanguage
  if payload.isActive is not None:
    user.is_active = payload.isActive
  if payload.permissions is not None:
    user.permissions = _normalize_permissions(payload.permissions, role=(payload.role or user.role), defaults_by_role=role_defaults)

  password_out = None
  if payload.password is not None:
    pw = (payload.password or '').strip()
    if not pw:
      raise HTTPException(status_code=400, detail='Пароль не должен быть пустым')
    user.temporary_password = pw
    user.password_hash = hash_password(pw)
    password_out = pw

  await db.commit()
  out = await db.execute(select(User).options(selectinload(User.company)).where(User.id == user.id))
  user = out.scalar_one()

  if password_out is None and current_user.role == UserRole.admin:
    password_out = user.temporary_password
  return TeamUserDetailsResponse(user=to_team_user_item(user), password=password_out)


@router.post(
  '/users/{user_id}/password/reset',
  response_model=TeamResetPasswordResponse,
  summary='Сбросить пароль участнику (временный)',
)
async def reset_user_password (
  user_id: UUID = Path(..., description='ID пользователя'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> TeamResetPasswordResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  res = await db.execute(select(User).where(User.id == user_id, User.company_id == company_id))
  user = res.scalar_one_or_none()
  if user is None:
    raise HTTPException(status_code=404, detail='Пользователь не найден')

  _ensure_can_manage_user(current_user, target_role=user.role)

  temporary_password = secrets.token_urlsafe(9)[:20]
  user.temporary_password = temporary_password
  user.password_hash = hash_password(temporary_password)
  await db.commit()
  return TeamResetPasswordResponse(password=temporary_password)


@router.get(
  '/users/{user_id}/project-access',
  response_model=UserProjectAccessResponse,
  summary='Доступ пользователя к проектам/регионам',
)
async def get_user_project_access (
  user_id: UUID = Path(..., description='ID пользователя'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserProjectAccessResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  await _ensure_target_user_in_company(db, company_id=company_id, user_id=user_id)

  # Managers are allowed to view only scopes inside their projects.
  stmt = (
    select(UserProjectAccess, Project.name)
    .join(Project, Project.id == UserProjectAccess.project_id)
    .where(UserProjectAccess.user_id == user_id, Project.company_id == company_id)
  )
  if (current_user.platform_role or 'user') != 'owner' and current_user.role == UserRole.manager:
    stmt = stmt.where(Project.manager_user_id == current_user.id)

  res = await db.execute(stmt)
  items = []
  for row, project_name in res.all():
    items.append(
      {
        'projectId': row.project_id,
        'projectName': project_name,
        'accessRole': row.access_role,
        'regionCodes': list(row.region_codes or []),
      },
    )
  return UserProjectAccessResponse(userId=user_id, items=items)


@router.put(
  '/users/{user_id}/project-access',
  response_model=UserProjectAccessResponse,
  summary='Заменить доступ пользователя к проектам/регионам',
)
async def replace_user_project_access (
  payload: UserProjectAccessReplaceRequest,
  user_id: UUID = Path(..., description='ID пользователя'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserProjectAccessResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  target = await _ensure_target_user_in_company(db, company_id=company_id, user_id=user_id)

  normalized = []
  for item in payload.items:
    role = _normalize_access_role(item.accessRole)
    regions = _normalize_region_codes(item.regionCodes)
    normalized.append((item.projectId, role, regions))

  # Validate permissions: for manager, every project must be managed by them.
  for project_id, _, _ in normalized:
    await _ensure_project_manage_access(db, actor=current_user, project_id=project_id)

  # Replace only in allowed scope set for manager; owner/admin can replace all.
  delete_stmt = select(UserProjectAccess.id).join(Project, Project.id == UserProjectAccess.project_id).where(
    UserProjectAccess.user_id == user_id,
    Project.company_id == company_id,
  )
  if (current_user.platform_role or 'user') != 'owner' and current_user.role == UserRole.manager:
    delete_stmt = delete_stmt.where(Project.manager_user_id == current_user.id)

  existing_ids_res = await db.execute(delete_stmt)
  existing_ids = [x for x in existing_ids_res.scalars().all()]
  if existing_ids:
    await db.execute(
      UserProjectAccess.__table__.delete().where(UserProjectAccess.id.in_(existing_ids)),  # type: ignore
    )

  for project_id, role, regions in normalized:
    db.add(
      UserProjectAccess(
        user_id=target.id,
        project_id=project_id,
        access_role=role,
        region_codes=regions,
        assigned_by_user_id=current_user.id,
      ),
    )

  await db.commit()
  return await get_user_project_access(user_id=user_id, db=db, current_user=current_user)


@router.post(
  '/users/project-access/bulk',
  response_model=BulkAssignResponse,
  summary='Массово назначить доступ (проект+регионы) пользователям',
)
async def bulk_assign_project_access (
  payload: BulkAssignRequest,
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> BulkAssignResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  await _ensure_project_manage_access(db, actor=current_user, project_id=payload.projectId)

  role = _normalize_access_role(payload.accessRole)
  regions = _normalize_region_codes(payload.regionCodes)

  user_ids = list(dict.fromkeys(payload.userIds))
  if not user_ids:
    return BulkAssignResponse(updatedUsers=0)

  # ensure all users belong to current company
  res = await db.execute(select(User.id).where(User.company_id == company_id, User.id.in_(user_ids)))
  allowed = set(res.scalars().all())
  if not allowed:
    raise HTTPException(status_code=400, detail='Пользователи не найдены')

  # upsert: delete existing for this project/user and insert
  await db.execute(
    UserProjectAccess.__table__.delete().where(  # type: ignore
      UserProjectAccess.project_id == payload.projectId,
      UserProjectAccess.user_id.in_(list(allowed)),
    ),
  )

  for uid in allowed:
    db.add(
      UserProjectAccess(
        user_id=uid,
        project_id=payload.projectId,
        access_role=role,
        region_codes=regions,
        assigned_by_user_id=current_user.id,
      ),
    )

  await db.commit()
  return BulkAssignResponse(updatedUsers=len(allowed))


@router.get(
  '/users/{user_id}/point-access',
  response_model=UserPointAccessResponse,
  summary='Привязки пользователя к локациям (точкам АП)',
)
async def get_user_point_access (
  user_id: UUID = Path(..., description='ID пользователя'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserPointAccessResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  company_id = _company_id(current_user)
  await _ensure_target_user_in_company(db, company_id=company_id, user_id=user_id)

  base_project_id = await _get_base_ap_project_id(db, company_id=company_id)
  if base_project_id is None:
    return UserPointAccessResponse(userId=user_id, projectId=None, pointIds=[])

  stmt = (
    select(UserPointAccess.point_id)
    .join(ShopPoint, ShopPoint.id == UserPointAccess.point_id)
    .where(UserPointAccess.user_id == user_id, ShopPoint.project_id == base_project_id)
  )
  res = await db.execute(stmt)
  return UserPointAccessResponse(userId=user_id, projectId=base_project_id, pointIds=list(res.scalars().all()))


@router.put(
  '/users/{user_id}/point-access',
  response_model=UserPointAccessResponse,
  summary='Заменить привязки пользователя к локациям (точкам АП)',
)
async def replace_user_point_access (
  payload: UserPointAccessReplaceRequest,
  user_id: UUID = Path(..., description='ID пользователя'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserPointAccessResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  company_id = _company_id(current_user)
  await _ensure_target_user_in_company(db, company_id=company_id, user_id=user_id)

  base_project_id = await _get_base_ap_project_id(db, company_id=company_id)
  if base_project_id is None:
    raise HTTPException(status_code=400, detail='Сначала загрузите адресную программу (АП)')

  # ensure base AP is within this company
  pres = await db.execute(select(Project.id).where(Project.id == base_project_id, Project.company_id == company_id))
  if pres.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Проект АП не найден')

  point_ids = list(dict.fromkeys(payload.pointIds))
  if point_ids:
    pres = await db.execute(
      select(ShopPoint.id).where(
        ShopPoint.project_id == base_project_id,
        ShopPoint.id.in_(point_ids),
      ),
    )
    allowed = set(pres.scalars().all())
    if len(allowed) != len(point_ids):
      raise HTTPException(status_code=400, detail='Некоторые точки не найдены в АП')
  else:
    allowed = set()

  # delete only bindings for points from base AP
  base_point_ids = select(ShopPoint.id).where(ShopPoint.project_id == base_project_id)
  await db.execute(
    delete(UserPointAccess).where(
      UserPointAccess.user_id == user_id,
      UserPointAccess.point_id.in_(base_point_ids),
    ),
  )

  for pid in allowed:
    db.add(UserPointAccess(user_id=user_id, point_id=pid, assigned_by_user_id=current_user.id))

  await db.commit()
  return await get_user_point_access(user_id=user_id, db=db, current_user=current_user)


def _normalize_header (value: str) -> str:
  return ''.join(ch for ch in str(value or '').strip().lower() if ch.isalnum() or ch in ('_', ' ')).replace('  ', ' ').strip()


def _normalize_group_name (value: str) -> str:
  return ' '.join(str(value or '').strip().split())


def _split_fio (fio: str) -> tuple[str, str]:
  parts = [p for p in str(fio or '').strip().split() if p]
  if len(parts) >= 2:
    return parts[1], parts[0]
  if len(parts) == 1:
    return parts[0], '—'
  return '—', '—'


@router.post(
  '/users/import',
  response_model=TeamUsersImportResponse,
  summary='Импорт участников из Excel (.xlsx)',
)
async def import_users_xlsx (
  file: UploadFile = File(...),
  update_existing: bool = Query(default=False, description='Обновлять существующих пользователей (по email)'),
  generate_password_if_empty: bool = Query(default=True, description='Генерировать пароль, если Password пустой'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> TeamUsersImportResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'data.import')

  if not file.filename or not file.filename.lower().endswith('.xlsx'):
    raise HTTPException(status_code=400, detail='Нужен файл .xlsx')

  raw = await file.read()
  if not raw:
    raise HTTPException(status_code=400, detail='Файл пустой')

  try:
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
  except Exception:
    raise HTTPException(status_code=400, detail='Не удалось прочитать .xlsx')

  rows = list(ws.iter_rows(values_only=True))
  if not rows or not rows[0]:
    raise HTTPException(status_code=400, detail='В файле нет заголовка')

  header = [_normalize_header(x) for x in rows[0]]
  idx: dict[str, int] = {h: i for i, h in enumerate(header) if h}

  def col (*names: str) -> int | None:
    for n in names:
      k = _normalize_header(n)
      if k in idx:
        return idx[k]
    return None

  i_email = col('email', 'электронная почта', 'почта')
  if i_email is None:
    raise HTTPException(status_code=400, detail='Не найдена колонка Email')

  i_password = col('password', 'пароль')
  i_fio = col('fio', 'фио')
  i_company = col('company', 'компания')
  i_phone = col('phone', 'телефон')
  i_role = col('role', 'роль')
  i_lang = col('language', 'язык', 'язык интерфейса')
  i_note = col('note', 'заметка')
  i_active = col('active', 'активен')

  errors: list[TeamUsersImportError] = []
  items: list[dict] = []

  for r_idx, row in enumerate(rows[1:], start=2):
    email = (str(row[i_email] or '').strip() if i_email is not None else '').lower()
    if not email:
      errors.append(TeamUsersImportError(row=r_idx, message='Email обязателен'))
      continue

    fio = str(row[i_fio] or '').strip() if i_fio is not None else ''
    first_name = ''
    last_name = ''
    if fio:
      first_name, last_name = _split_fio(fio)

    profile_company = str(row[i_company] or '').strip() if i_company is not None else ''
    phone = str(row[i_phone] or '').strip() if i_phone is not None else ''
    role_raw = str(row[i_role] or '').strip().lower() if i_role is not None else ''
    lang = str(row[i_lang] or '').strip().lower() if i_lang is not None else ''
    note = str(row[i_note] or '').strip() if i_note is not None else ''
    active_raw = str(row[i_active] or '').strip().lower() if i_active is not None else ''

    password = str(row[i_password] or '').strip() if i_password is not None else ''
    if not password and generate_password_if_empty:
      password = secrets.token_urlsafe(9)[:20]

    is_active = True
    if active_raw in ('0', 'false', 'no', 'off', 'нет'):
      is_active = False
    if active_raw in ('1', 'true', 'yes', 'on', 'да'):
      is_active = True

    if role_raw and role_raw not in {r.value for r in UserRole}:
      errors.append(TeamUsersImportError(row=r_idx, message='Некорректная роль'))
      continue

    role = UserRole(role_raw) if role_raw else UserRole.manager

    if not first_name or first_name == '—':
      first_name = '—'
    if not last_name or last_name == '—':
      last_name = '—'

    if not lang:
      lang = 'ru'

    items.append(
      {
        'email': email,
        'password': password or None,
        'first_name': first_name,
        'last_name': last_name,
        'phone': phone or None,
        'role': role,
        'profile_company': profile_company or None,
        'ui_language': lang,
        'note': note or None,
        'is_active': is_active,
        'row': r_idx,
      },
    )

  if errors:
    return TeamUsersImportResponse(created=0, updated=0, skipped=0, errors=errors)

  company_id = _company_id(current_user)

  created = 0
  updated = 0
  skipped = 0

  for item in items:
    res = await db.execute(select(User).where(User.company_id == company_id, User.email == item['email']))
    existing = res.scalar_one_or_none()

    if existing is None:
      _ensure_can_manage_user(current_user, target_role=item['role'])
      pw = item['password'] or secrets.token_urlsafe(9)[:20]
      user = User(
        first_name=item['first_name'],
        last_name=item['last_name'],
        email=item['email'],
        phone=item['phone'],
        role=item['role'],
        profile_company=item['profile_company'],
        ui_language=item['ui_language'],
        is_active=bool(item['is_active']),
        note=item['note'],
        company_id=company_id,
        temporary_password=pw,
        password_hash=hash_password(pw),
      )
      db.add(user)
      created += 1
      continue

    if not update_existing:
      skipped += 1
      continue

    _ensure_can_manage_user(current_user, target_role=existing.role)
    _ensure_can_manage_user(current_user, target_role=item['role'])

    existing.first_name = item['first_name']
    existing.last_name = item['last_name']
    existing.phone = item['phone']
    existing.role = item['role']
    existing.profile_company = item['profile_company']
    existing.ui_language = item['ui_language']
    existing.is_active = bool(item['is_active'])
    existing.note = item['note']

    if item['password']:
      existing.temporary_password = item['password']
      existing.password_hash = hash_password(item['password'])
    updated += 1

  try:
    await db.commit()
  except Exception as err:
    await db.rollback()
    return TeamUsersImportResponse(created=0, updated=0, skipped=0, errors=[TeamUsersImportError(row=0, message=str(err))])

  return TeamUsersImportResponse(created=created, updated=updated, skipped=skipped, errors=None)


async def _get_group_or_404 (db: AsyncSession, *, company_id: UUID, group_id: UUID) -> UserGroup:
  res = await db.execute(select(UserGroup).where(UserGroup.id == group_id, UserGroup.company_id == company_id))
  group = res.scalar_one_or_none()
  if group is None:
    raise HTTPException(status_code=404, detail='Группа не найдена')
  return group


@router.get('/groups', response_model=UserGroupsResponse, summary='Список пользовательских групп')
async def list_user_groups (
  q: str | None = Query(default=None, description='Поиск по названию группы'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserGroupsResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)

  stmt = select(UserGroup).where(UserGroup.company_id == company_id).order_by(UserGroup.created_at.desc())
  if q:
    stmt = stmt.where(UserGroup.name.ilike(f'%{q.strip()}%'))
  res = await db.execute(stmt)
  groups = res.scalars().all()

  # members count
  ids = [g.id for g in groups]
  counts: dict[UUID, int] = {}
  if ids:
    cnt_res = await db.execute(
      select(UserGroupMember.group_id, func.count(UserGroupMember.id))
      .where(UserGroupMember.group_id.in_(ids))
      .group_by(UserGroupMember.group_id),
    )
    counts = {gid: int(c) for gid, c in cnt_res.all()}

  items = [UserGroupItem(id=g.id, name=g.name, membersCount=counts.get(g.id, 0), createdAt=g.created_at) for g in groups]
  return UserGroupsResponse(items=items)


@router.post('/groups', response_model=UserGroupItem, status_code=201, summary='Создать группу')
async def create_user_group (
  payload: CreateUserGroupRequest,
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserGroupItem:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)

  name = _normalize_group_name(payload.name)
  if not name:
    raise HTTPException(status_code=400, detail='Название группы пустое')

  exists = await db.execute(
    select(UserGroup.id).where(UserGroup.company_id == company_id, func.lower(UserGroup.name) == name.lower()),
  )
  if exists.scalar_one_or_none() is not None:
    raise HTTPException(status_code=400, detail='Группа с таким названием уже существует')

  group = UserGroup(company_id=company_id, name=name, created_by_user_id=current_user.id)
  db.add(group)
  await db.commit()
  return UserGroupItem(id=group.id, name=group.name, membersCount=0, createdAt=group.created_at)


@router.patch('/groups/{group_id}', response_model=UserGroupItem, summary='Переименовать группу')
async def update_user_group (
  payload: UpdateUserGroupRequest,
  group_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserGroupItem:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  group = await _get_group_or_404(db, company_id=company_id, group_id=group_id)

  name = _normalize_group_name(payload.name)
  if not name:
    raise HTTPException(status_code=400, detail='Название группы пустое')

  exists = await db.execute(
    select(UserGroup.id).where(
      UserGroup.company_id == company_id,
      func.lower(UserGroup.name) == name.lower(),
      UserGroup.id != group.id,
    ),
  )
  if exists.scalar_one_or_none() is not None:
    raise HTTPException(status_code=400, detail='Группа с таким названием уже существует')

  group.name = name
  await db.commit()

  cnt = await db.execute(select(func.count()).select_from(UserGroupMember).where(UserGroupMember.group_id == group.id))
  return UserGroupItem(id=group.id, name=group.name, membersCount=int(cnt.scalar_one() or 0), createdAt=group.created_at)


@router.delete('/groups/{group_id}', response_model=dict, summary='Удалить группу')
async def delete_user_group (
  group_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> dict:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  group = await _get_group_or_404(db, company_id=company_id, group_id=group_id)

  await db.execute(delete(UserGroup).where(UserGroup.id == group.id))
  await db.commit()
  return {'ok': True}


@router.get('/groups/{group_id}/members', response_model=UserGroupMembersResponse, summary='Участники группы')
async def get_group_members (
  group_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserGroupMembersResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  await _get_group_or_404(db, company_id=company_id, group_id=group_id)

  res = await db.execute(
    select(UserGroupMember.user_id).join(User, User.id == UserGroupMember.user_id).where(
      UserGroupMember.group_id == group_id,
      User.company_id == company_id,
    ),
  )
  return UserGroupMembersResponse(groupId=group_id, userIds=list(res.scalars().all()))


@router.put('/groups/{group_id}/members', response_model=UserGroupMembersResponse, summary='Заменить участников группы')
async def replace_group_members (
  payload: ReplaceUserGroupMembersRequest,
  group_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserGroupMembersResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  await _get_group_or_404(db, company_id=company_id, group_id=group_id)

  user_ids = list(dict.fromkeys(payload.userIds))
  if user_ids:
    allowed = await db.execute(select(User.id).where(User.company_id == company_id, User.id.in_(user_ids)))
    allowed_set = set(allowed.scalars().all())
    if len(allowed_set) != len(user_ids):
      raise HTTPException(status_code=400, detail='Некоторые пользователи не найдены в компании')
  else:
    allowed_set = set()

  await db.execute(delete(UserGroupMember).where(UserGroupMember.group_id == group_id))
  for uid in allowed_set:
    db.add(UserGroupMember(group_id=group_id, user_id=uid))
  await db.commit()
  return await get_group_members(group_id=group_id, db=db, current_user=current_user)


@router.post('/groups/{group_id}/members/add-by-email', response_model=UserGroupMembersResponse, summary='Добавить участника по email')
async def add_group_member_by_email (
  payload: AddUserGroupMemberByEmailRequest,
  group_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserGroupMembersResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  await _get_group_or_404(db, company_id=company_id, group_id=group_id)

  res = await db.execute(select(User).where(User.company_id == company_id, func.lower(User.email) == str(payload.email).lower()))
  user = res.scalar_one_or_none()
  if user is None:
    raise HTTPException(status_code=404, detail='Пользователь не найден')

  db.add(UserGroupMember(group_id=group_id, user_id=user.id))
  try:
    await db.commit()
  except Exception:
    await db.rollback()
  return await get_group_members(group_id=group_id, db=db, current_user=current_user)


@router.post(
  '/groups/{group_id}/members/import',
  response_model=UserGroupMembersImportResponse,
  summary='Импорт участников группы из Excel (.xlsx)',
)
async def import_group_members_xlsx (
  group_id: UUID = Path(...),
  file: UploadFile = File(...),
  mode: str = Query(default='add', description='add|replace'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserGroupMembersImportResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'data.import')
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  await _get_group_or_404(db, company_id=company_id, group_id=group_id)

  if not file.filename or not file.filename.lower().endswith('.xlsx'):
    raise HTTPException(status_code=400, detail='Нужен файл .xlsx')
  raw = await file.read()
  if not raw:
    raise HTTPException(status_code=400, detail='Файл пустой')

  try:
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
  except Exception:
    raise HTTPException(status_code=400, detail='Не удалось прочитать .xlsx')

  rows = list(ws.iter_rows(values_only=True))
  if not rows or not rows[0]:
    raise HTTPException(status_code=400, detail='В файле нет заголовка')

  header = [_normalize_header(x) for x in rows[0]]
  idx = {h: i for i, h in enumerate(header) if h}
  if 'email' not in idx and 'почта' not in idx and 'e-mail' not in idx:
    raise HTTPException(status_code=400, detail='Не найдена колонка Email')
  i_email = idx.get('email', idx.get('почта', idx.get('e-mail')))

  emails: list[tuple[int, str]] = []
  errors: list[UserGroupMembersImportError] = []
  for r_idx, row in enumerate(rows[1:], start=2):
    email = str(row[i_email] or '').strip().lower()
    if not email:
      errors.append(UserGroupMembersImportError(row=r_idx, message='Email обязателен'))
      continue
    emails.append((r_idx, email))

  if errors:
    return UserGroupMembersImportResponse(added=0, skipped=0, errors=errors)

  if mode not in ('add', 'replace'):
    raise HTTPException(status_code=400, detail='Некорректный mode')

  if mode == 'replace':
    await db.execute(delete(UserGroupMember).where(UserGroupMember.group_id == group_id))

  added = 0
  skipped = 0
  for r_idx, email in emails:
    res = await db.execute(select(User.id).where(User.company_id == company_id, func.lower(User.email) == email))
    uid = res.scalar_one_or_none()
    if uid is None:
      errors.append(UserGroupMembersImportError(row=r_idx, message='Пользователь не найден'))
      continue
    db.add(UserGroupMember(group_id=group_id, user_id=uid))
    try:
      await db.flush()
      added += 1
    except Exception:
      await db.rollback()
      skipped += 1

  await db.commit()
  return UserGroupMembersImportResponse(added=added, skipped=skipped, errors=errors or None)


@router.get('/users/{user_id}/groups', response_model=UserGroupsForUserResponse, summary='Группы пользователя')
async def get_user_groups (
  user_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserGroupsForUserResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  await _ensure_target_user_in_company(db, company_id=company_id, user_id=user_id)

  res = await db.execute(
    select(UserGroupMember.group_id)
    .join(UserGroup, UserGroup.id == UserGroupMember.group_id)
    .where(UserGroupMember.user_id == user_id, UserGroup.company_id == company_id),
  )
  return UserGroupsForUserResponse(userId=user_id, groupIds=list(res.scalars().all()))


@router.put('/users/{user_id}/groups', response_model=UserGroupsForUserResponse, summary='Заменить группы пользователя')
async def replace_user_groups (
  payload: ReplaceUserGroupsForUserRequest,
  user_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserGroupsForUserResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  company_id = _company_id(current_user)
  await _ensure_target_user_in_company(db, company_id=company_id, user_id=user_id)

  group_ids = list(dict.fromkeys(payload.groupIds))
  if group_ids:
    allowed = await db.execute(select(UserGroup.id).where(UserGroup.company_id == company_id, UserGroup.id.in_(group_ids)))
    allowed_set = set(allowed.scalars().all())
    if len(allowed_set) != len(group_ids):
      raise HTTPException(status_code=400, detail='Некоторые группы не найдены в компании')
  else:
    allowed_set = set()

  # delete all memberships for this user inside company
  await db.execute(
    delete(UserGroupMember).where(
      UserGroupMember.user_id == user_id,
      UserGroupMember.group_id.in_(select(UserGroup.id).where(UserGroup.company_id == company_id)),
    ),
  )
  for gid in allowed_set:
    db.add(UserGroupMember(group_id=gid, user_id=user_id))
  await db.commit()

  return await get_user_groups(user_id=user_id, db=db, current_user=current_user)


async def _list_manageable_companies (db: AsyncSession, *, actor: User) -> list[tuple[Company, CompanySettings | None]]:
  # owner: all companies they have access to
  if (actor.platform_role or 'user') == 'owner':
    stmt = (
      select(Company, CompanySettings)
      .join(OwnerCompanyAccess, OwnerCompanyAccess.company_id == Company.id)
      .outerjoin(CompanySettings, CompanySettings.company_id == Company.id)
      .where(OwnerCompanyAccess.owner_user_id == actor.id)
      .order_by(Company.created_at.desc())
    )
    res = await db.execute(stmt)
    return list(res.all())

  # non-owner: only their active company
  cid = _company_id(actor)
  stmt = select(Company, CompanySettings).outerjoin(CompanySettings, CompanySettings.company_id == Company.id).where(Company.id == cid)
  res = await db.execute(stmt)
  return list(res.all())


@router.get('/users/{user_id}/companies-access', response_model=UserCompaniesAccessResponse, summary='Доступ пользователя к клиентам')
async def get_user_companies_access (
  user_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserCompaniesAccessResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  home_company_id = _company_id(current_user)
  if home_company_id is None:
    raise HTTPException(status_code=400, detail='Не выбран активный клиент')
  # manage only users from current company (home workspace)
  await _ensure_target_user_in_company(db, company_id=home_company_id, user_id=user_id)

  companies = await _list_manageable_companies(db, actor=current_user)
  ids = [c.id for c, _ in companies]

  granted: set[UUID] = set()
  if ids:
    res = await db.execute(
      select(UserCompanyAccess.company_id).where(UserCompanyAccess.user_id == user_id, UserCompanyAccess.company_id.in_(ids)),
    )
    granted = set(res.scalars().all())

  items = []
  for c, s in companies:
    items.append(
      DistributionCompanyItem(
        id=c.id,
        name=c.name,
        baseApProjectId=(s.base_ap_project_id if s is not None else None),
        hasAccess=c.id in granted or c.id == home_company_id,
      ),
    )
  return UserCompaniesAccessResponse(userId=user_id, items=items)


@router.put('/users/{user_id}/companies-access', response_model=UserCompaniesAccessResponse, summary='Заменить доступ пользователя к клиентам')
async def replace_user_companies_access (
  payload: ReplaceUserCompaniesAccessRequest,
  user_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserCompaniesAccessResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  home_company_id = _company_id(current_user)
  if home_company_id is None:
    raise HTTPException(status_code=400, detail='Не выбран активный клиент')
  await _ensure_target_user_in_company(db, company_id=home_company_id, user_id=user_id)

  companies = await _list_manageable_companies(db, actor=current_user)
  allowed_ids = {c.id for c, _ in companies}

  next_ids = set(payload.companyIds)
  # always keep home company
  next_ids.add(home_company_id)

  if not next_ids.issubset(allowed_ids):
    raise HTTPException(status_code=400, detail='Некоторые клиенты недоступны для назначения')

  # replace: delete all grants in allowed scope except home
  await db.execute(
    delete(UserCompanyAccess).where(
      UserCompanyAccess.user_id == user_id,
      UserCompanyAccess.company_id.in_(list(allowed_ids - {home_company_id})),
    ),
  )
  for cid in sorted(next_ids):
    if cid == home_company_id:
      continue
    db.add(UserCompanyAccess(user_id=user_id, company_id=cid, created_by_user_id=current_user.id))

  await db.commit()
  return await get_user_companies_access(user_id=user_id, db=db, current_user=current_user)


@router.get('/users/{user_id}/distribution/{company_id}', response_model=UserCompanyDistributionResponse, summary='Распределение пользователя по клиенту')
async def get_user_company_distribution (
  user_id: UUID = Path(...),
  company_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserCompanyDistributionResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  home_company_id = _company_id(current_user)
  await _ensure_target_user_in_company(db, company_id=home_company_id, user_id=user_id)

  # require that actor can manage this company
  manageable = await _list_manageable_companies(db, actor=current_user)
  allowed_company_ids = {c.id for c, _ in manageable}
  if company_id not in allowed_company_ids:
    raise HTTPException(status_code=403, detail='Нет доступа к клиенту')

  # require that user has access to this company (or home)
  if company_id != home_company_id:
    res = await db.execute(select(UserCompanyAccess.id).where(UserCompanyAccess.user_id == user_id, UserCompanyAccess.company_id == company_id))
    if res.scalar_one_or_none() is None:
      return UserCompanyDistributionResponse(userId=user_id, companyId=company_id, filterValues={}, pointIds=[])

  # filters
  dist_res = await db.execute(select(UserCompanyDistribution).where(UserCompanyDistribution.user_id == user_id, UserCompanyDistribution.company_id == company_id))
  dist = dist_res.scalar_one_or_none()
  filter_values = dist.filter_values if dist is not None else {}

  # points from base AP
  settings = await db.execute(select(CompanySettings).where(CompanySettings.company_id == company_id))
  s = settings.scalar_one_or_none()
  project_id = s.base_ap_project_id if s is not None else None
  point_ids: list[UUID] = []
  if project_id is not None:
    pres = await db.execute(select(UserPointAccess.point_id).join(ShopPoint, ShopPoint.id == UserPointAccess.point_id).where(UserPointAccess.user_id == user_id, ShopPoint.project_id == project_id))
    point_ids = list(pres.scalars().all())

  return UserCompanyDistributionResponse(userId=user_id, companyId=company_id, filterValues=filter_values or {}, pointIds=point_ids)


@router.put('/users/{user_id}/distribution/{company_id}', response_model=UserCompanyDistributionResponse, summary='Сохранить распределение пользователя по клиенту')
async def replace_user_company_distribution (
  payload: ReplaceUserCompanyDistributionRequest,
  user_id: UUID = Path(...),
  company_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserCompanyDistributionResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  home_company_id = _company_id(current_user)
  await _ensure_target_user_in_company(db, company_id=home_company_id, user_id=user_id)

  manageable = await _list_manageable_companies(db, actor=current_user)
  allowed_company_ids = {c.id for c, _ in manageable}
  if company_id not in allowed_company_ids:
    raise HTTPException(status_code=403, detail='Нет доступа к клиенту')

  # ensure user has access (or home)
  if company_id != home_company_id:
    res = await db.execute(select(UserCompanyAccess.id).where(UserCompanyAccess.user_id == user_id, UserCompanyAccess.company_id == company_id))
    if res.scalar_one_or_none() is None:
      raise HTTPException(status_code=400, detail='Сначала выдайте доступ пользователю к клиенту')

  # upsert distribution
  dist_res = await db.execute(select(UserCompanyDistribution).where(UserCompanyDistribution.user_id == user_id, UserCompanyDistribution.company_id == company_id))
  dist = dist_res.scalar_one_or_none()
  if dist is None:
    dist = UserCompanyDistribution(user_id=user_id, company_id=company_id, filter_values=payload.filterValues or {}, updated_by_user_id=current_user.id)
    db.add(dist)
  else:
    dist.filter_values = payload.filterValues or {}
    dist.updated_by_user_id = current_user.id

  # replace point bindings only within base AP of that company
  settings_res = await db.execute(select(CompanySettings).where(CompanySettings.company_id == company_id))
  s = settings_res.scalar_one_or_none()
  project_id = s.base_ap_project_id if s is not None else None
  if project_id is not None:
    allowed_points = set()
    pids = list(dict.fromkeys(payload.pointIds or []))
    if pids:
      allowed_res = await db.execute(select(ShopPoint.id).where(ShopPoint.project_id == project_id, ShopPoint.id.in_(pids)))
      allowed_points = set(allowed_res.scalars().all())
      if len(allowed_points) != len(pids):
        raise HTTPException(status_code=400, detail='Некоторые точки не найдены в АП')

    base_point_ids = select(ShopPoint.id).where(ShopPoint.project_id == project_id)
    await db.execute(delete(UserPointAccess).where(UserPointAccess.user_id == user_id, UserPointAccess.point_id.in_(base_point_ids)))
    for pid in allowed_points:
      db.add(UserPointAccess(user_id=user_id, point_id=pid, assigned_by_user_id=current_user.id))

  await db.commit()
  return await get_user_company_distribution(user_id=user_id, company_id=company_id, db=db, current_user=current_user)


@router.get('/distribution/{company_id}/filter-values', response_model=CompanyFilterValuesResponse, summary='Доступные значения фильтров клиента (по АП)')
async def get_company_filter_values (
  company_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> CompanyFilterValuesResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')

  manageable = await _list_manageable_companies(db, actor=current_user)
  allowed_company_ids = {c.id for c, _ in manageable}
  if company_id not in allowed_company_ids:
    raise HTTPException(status_code=403, detail='Нет доступа к клиенту')

  settings_res = await db.execute(select(CompanySettings).where(CompanySettings.company_id == company_id))
  s = settings_res.scalar_one_or_none()
  if s is None or s.base_ap_project_id is None:
    return CompanyFilterValuesResponse(companyId=company_id, items=[])

  filters_def = list(s.filters or [])
  keys = []
  titles_by_key: dict[str, str] = {}
  for f in filters_def:
    if isinstance(f, dict) and isinstance(f.get('key'), str):
      keys.append(f['key'])
      if isinstance(f.get('title'), str) and f.get('title'):
        titles_by_key[f['key']] = f['title']

  if not keys:
    return CompanyFilterValuesResponse(companyId=company_id, items=[])

  # load points and collect unique values
  res = await db.execute(select(ShopPoint).where(ShopPoint.project_id == s.base_ap_project_id))
  points = res.scalars().all()

  out = []
  for key in keys:
    seen: set[str] = set()
    vals: list[str] = []
    for p in points:
      v = None
      if key == 'region':
        v = p.region_code
      elif key == 'city':
        v = p.city_name
      else:
        raw = (p.attrs or {}).get(str(key))
        if isinstance(raw, str):
          v = raw
      if not v:
        continue
      s_v = str(v).strip()
      if not s_v or s_v in seen:
        continue
      seen.add(s_v)
      vals.append(s_v)
      if len(vals) >= 250:
        break
    out.append({'key': key, 'title': titles_by_key.get(key), 'values': sorted(vals)})

  return CompanyFilterValuesResponse(companyId=company_id, items=out)


@router.get('/users/{user_id}/reports/{company_id}', response_model=UserCompanyReportsResponse, summary='Доступные отчёты пользователя по клиенту')
async def get_user_company_reports (
  user_id: UUID = Path(...),
  company_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserCompanyReportsResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  home_company_id = _company_id(current_user)
  await _ensure_target_user_in_company(db, company_id=home_company_id, user_id=user_id)

  manageable = await _list_manageable_companies(db, actor=current_user)
  allowed_company_ids = {c.id for c, _ in manageable}
  if company_id not in allowed_company_ids:
    raise HTTPException(status_code=403, detail='Нет доступа к клиенту')

  if company_id != home_company_id:
    res = await db.execute(select(UserCompanyAccess.id).where(UserCompanyAccess.user_id == user_id, UserCompanyAccess.company_id == company_id))
    if res.scalar_one_or_none() is None:
      return UserCompanyReportsResponse(userId=user_id, companyId=company_id, reportKeys=[])

  res = await db.execute(select(UserCompanyReports).where(UserCompanyReports.user_id == user_id, UserCompanyReports.company_id == company_id))
  row = res.scalar_one_or_none()
  return UserCompanyReportsResponse(userId=user_id, companyId=company_id, reportKeys=(row.report_keys if row is not None else []))


@router.put('/users/{user_id}/reports/{company_id}', response_model=UserCompanyReportsResponse, summary='Сохранить доступные отчёты пользователя по клиенту')
async def replace_user_company_reports (
  payload: ReplaceUserCompanyReportsRequest,
  user_id: UUID = Path(...),
  company_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> UserCompanyReportsResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')
  home_company_id = _company_id(current_user)
  await _ensure_target_user_in_company(db, company_id=home_company_id, user_id=user_id)

  manageable = await _list_manageable_companies(db, actor=current_user)
  allowed_company_ids = {c.id for c, _ in manageable}
  if company_id not in allowed_company_ids:
    raise HTTPException(status_code=403, detail='Нет доступа к клиенту')

  if company_id != home_company_id:
    res = await db.execute(select(UserCompanyAccess.id).where(UserCompanyAccess.user_id == user_id, UserCompanyAccess.company_id == company_id))
    if res.scalar_one_or_none() is None:
      raise HTTPException(status_code=400, detail='Сначала выдайте доступ пользователю к клиенту')

  keys = []
  seen = set()
  for k in payload.reportKeys or []:
    v = str(k or '').strip()
    if not v or v in seen:
      continue
    seen.add(v)
    keys.append(v[:80])
    if len(keys) >= 300:
      break

  res = await db.execute(select(UserCompanyReports).where(UserCompanyReports.user_id == user_id, UserCompanyReports.company_id == company_id))
  row = res.scalar_one_or_none()
  if row is None:
    row = UserCompanyReports(user_id=user_id, company_id=company_id, report_keys=keys, updated_by_user_id=current_user.id)
    db.add(row)
  else:
    row.report_keys = keys
    row.updated_by_user_id = current_user.id

  await db.commit()
  return await get_user_company_reports(user_id=user_id, company_id=company_id, db=db, current_user=current_user)


@router.post('/users/{user_id}/clone', response_model=CloneUserResponse, summary='Клонировать настройки пользователя')
async def clone_user_settings (
  payload: CloneUserRequest,
  user_id: UUID = Path(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> CloneUserResponse:
  _ensure_admin_or_manager_or_owner(current_user)
  _ensure_perm(current_user, 'users.edit')

  home_company_id = _company_id(current_user)
  target = await _ensure_target_user_in_company(db, company_id=home_company_id, user_id=user_id)
  source = await _ensure_target_user_in_company(db, company_id=home_company_id, user_id=payload.sourceUserId)

  _ensure_can_manage_user(current_user, target_role=target.role)
  _ensure_can_manage_user(current_user, target_role=source.role)

  if source.id == target.id:
    raise HTTPException(status_code=400, detail='Нельзя клонировать пользователя в самого себя')

  mode = payload.mode
  do_role = mode in ('all', 'role_groups')
  do_distribution = mode in ('all', 'distribution')
  do_reports = mode in ('all', 'reports')

  changed_role = False
  changed_groups = 0
  changed_companies = 0
  changed_company_distributions = 0
  changed_points = 0
  changed_company_reports = 0

  manageable = await _list_manageable_companies(db, actor=current_user)
  allowed_company_ids = {c.id for c, _ in manageable}
  allowed_company_ids.add(home_company_id)

  if do_role:
    if target.role != source.role or (target.permissions or []) != (source.permissions or []):
      target.role = source.role
      target.permissions = list(source.permissions or [])
      changed_role = True

    # group memberships only within home company
    g_res = await db.execute(
      select(UserGroupMember.group_id)
      .join(UserGroup, UserGroup.id == UserGroupMember.group_id)
      .where(UserGroupMember.user_id == source.id, UserGroup.company_id == home_company_id),
    )
    source_group_ids = set(g_res.scalars().all())

    await db.execute(
      delete(UserGroupMember)
      .where(UserGroupMember.user_id == target.id)
      .where(UserGroupMember.group_id.in_(select(UserGroup.id).where(UserGroup.company_id == home_company_id))),
    )
    for gid in sorted(source_group_ids):
      db.add(UserGroupMember(group_id=gid, user_id=target.id))
    changed_groups = len(source_group_ids)

  if do_distribution:
    # company access
    access_res = await db.execute(
      select(UserCompanyAccess.company_id).where(UserCompanyAccess.user_id == source.id, UserCompanyAccess.company_id.in_(list(allowed_company_ids))),
    )
    source_access_ids = set(access_res.scalars().all())
    source_access_ids.add(home_company_id)

    # replace target access within allowed scope (excluding home)
    await db.execute(
      delete(UserCompanyAccess).where(
        UserCompanyAccess.user_id == target.id,
        UserCompanyAccess.company_id.in_(list(allowed_company_ids - {home_company_id})),
      ),
    )
    for cid in sorted(source_access_ids):
      if cid == home_company_id:
        continue
      db.add(UserCompanyAccess(user_id=target.id, company_id=cid, created_by_user_id=current_user.id))
    changed_companies = len(source_access_ids)

    # company distributions (filter_values)
    dist_res = await db.execute(
      select(UserCompanyDistribution).where(
        UserCompanyDistribution.user_id == source.id,
        UserCompanyDistribution.company_id.in_(list(allowed_company_ids)),
      ),
    )
    source_dists = dist_res.scalars().all()

    await db.execute(
      delete(UserCompanyDistribution).where(
        UserCompanyDistribution.user_id == target.id,
        UserCompanyDistribution.company_id.in_(list(allowed_company_ids)),
      ),
    )
    for d in source_dists:
      db.add(
        UserCompanyDistribution(
          user_id=target.id,
          company_id=d.company_id,
          filter_values=d.filter_values or {},
          updated_by_user_id=current_user.id,
        ),
      )
    changed_company_distributions = len(source_dists)

    # points per company base AP project
    settings_res = await db.execute(select(CompanySettings).where(CompanySettings.company_id.in_(list(allowed_company_ids))))
    settings = settings_res.scalars().all()
    base_by_company: dict[UUID, UUID] = {}
    for s in settings:
      if s.base_ap_project_id is not None:
        base_by_company[s.company_id] = s.base_ap_project_id

    for company_id, project_id in base_by_company.items():
      # if source has no access and it's not home, skip cloning points
      if company_id != home_company_id and company_id not in source_access_ids:
        continue

      s_points_res = await db.execute(
        select(UserPointAccess.point_id)
        .join(ShopPoint, ShopPoint.id == UserPointAccess.point_id)
        .where(UserPointAccess.user_id == source.id, ShopPoint.project_id == project_id),
      )
      source_point_ids = set(s_points_res.scalars().all())

      base_point_ids_stmt = select(ShopPoint.id).where(ShopPoint.project_id == project_id)
      await db.execute(
        delete(UserPointAccess).where(UserPointAccess.user_id == target.id, UserPointAccess.point_id.in_(base_point_ids_stmt)),
      )
      for pid in sorted(source_point_ids):
        db.add(UserPointAccess(user_id=target.id, point_id=pid, assigned_by_user_id=current_user.id))
      changed_points += len(source_point_ids)

  if do_reports:
    rep_res = await db.execute(
      select(UserCompanyReports).where(
        UserCompanyReports.user_id == source.id,
        UserCompanyReports.company_id.in_(list(allowed_company_ids)),
      ),
    )
    source_reports = rep_res.scalars().all()

    await db.execute(
      delete(UserCompanyReports).where(
        UserCompanyReports.user_id == target.id,
        UserCompanyReports.company_id.in_(list(allowed_company_ids)),
      ),
    )
    for r in source_reports:
      db.add(
        UserCompanyReports(
          user_id=target.id,
          company_id=r.company_id,
          report_keys=list(r.report_keys or []),
          updated_by_user_id=current_user.id,
        ),
      )
    changed_company_reports = len(source_reports)

  await db.commit()
  return CloneUserResponse(
    targetUserId=target.id,
    sourceUserId=source.id,
    mode=mode,
    changedRole=changed_role,
    changedGroups=changed_groups,
    changedCompanies=changed_companies,
    changedCompanyDistributions=changed_company_distributions,
    changedPoints=changed_points,
    changedCompanyReports=changed_company_reports,
  )

