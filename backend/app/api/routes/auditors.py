import io
from datetime import date, datetime
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Path, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.core.security import hash_password
from app.db.session import get_db
from app.models.auditor import Auditor
from app.models.user import User
from app.models.user_role import UserRole
from app.schemas.auditors import (
  AuditorCreateRequest,
  AuditorCreateResponse,
  AuditorImportErrorItem,
  AuditorImportResponse,
  AuditorItem,
  AuditorSetPasswordRequest,
  AuditorUpdateRequest,
  AuditorUpdateResponse,
  AuditorsResponse,
  AuditorGender,
)


router = APIRouter()

def _company_id (current_user: User) -> UUID:
  return getattr(current_user, 'active_company_id', current_user.company_id)


def _ensure_admin (user: User) -> None:
  if user.role != UserRole.admin:
    raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail='Недостаточно прав')


def _norm_phone (value: str | None) -> str | None:
  s = (value or '').strip()
  return s if s else None


def _norm_email (value: str | None) -> str | None:
  s = (value or '').strip().lower()
  return s if s else None


def _parse_birth_date (value: object | None) -> date | None:
  if value is None:
    return None
  if isinstance(value, date) and not isinstance(value, datetime):
    return value
  if isinstance(value, datetime):
    return value.date()
  s = str(value).strip()
  if not s:
    return None
  # YYYY-MM-DD
  try:
    return date.fromisoformat(s)
  except Exception:
    pass
  # DD.MM.YYYY
  try:
    parts = s.split('.')
    if len(parts) == 3:
      dd, mm, yyyy = parts
      return date(int(yyyy), int(mm), int(dd))
  except Exception:
    pass
  # DD-MM-YYYY
  try:
    parts = s.split('-')
    if len(parts) == 3:
      p1, p2, p3 = parts
      # если первый блок — год, то это уже было обработано ISO выше
      if len(p1) != 4 and len(p3) == 4:
        dd, mm, yyyy = p1, p2, p3
        return date(int(yyyy), int(mm), int(dd))
  except Exception:
    pass
  return None


def _parse_gender (value: object | None) -> AuditorGender | None:
  s = str(value or '').strip().lower()
  if not s:
    return None
  if s in ('male', 'm', 'м', 'муж', 'мужской'):
    return 'male'
  if s in ('female', 'f', 'ж', 'жен', 'женский'):
    return 'female'
  return None


def to_auditor_item (a: Auditor) -> AuditorItem:
  return AuditorItem(
    id=a.id,
    lastName=a.last_name,
    firstName=a.first_name,
    middleName=a.middle_name,
    phone=a.phone,
    email=a.email,
    city=a.city,
    birthDate=a.birth_date,
    gender=(a.gender if a.gender in ('male', 'female') else None),
    createdAt=a.created_at,
  )


def _xlsx_bytes_from_rows (headers: list[str], rows: list[list[object | None]]) -> bytes:
  wb = Workbook(write_only=True)
  ws = wb.create_sheet(title='Sheet1')
  ws.append(headers)
  for row in rows:
    ws.append([('' if value is None else value) for value in row])

  buf = io.BytesIO()
  wb.save(buf)
  return buf.getvalue()


async def _exists_duplicate (
  db: AsyncSession,
  company_id: UUID,
  email: str | None,
  phone: str | None,
) -> bool:
  if email and phone:
    stmt = select(Auditor.id).where(Auditor.company_id == company_id, func.lower(Auditor.email) == email.lower(), Auditor.phone == phone)
  elif email:
    stmt = select(Auditor.id).where(Auditor.company_id == company_id, func.lower(Auditor.email) == email.lower())
  elif phone:
    stmt = select(Auditor.id).where(Auditor.company_id == company_id, Auditor.phone == phone)
  else:
    return False

  res = await db.execute(stmt.limit(1))
  return res.scalar_one_or_none() is not None


async def _exists_duplicate_excluding (
  db: AsyncSession,
  company_id: UUID,
  exclude_auditor_id: UUID,
  email: str | None,
  phone: str | None,
) -> bool:
  if email and phone:
    stmt = select(Auditor.id).where(
      Auditor.company_id == company_id,
      Auditor.id != exclude_auditor_id,
      func.lower(Auditor.email) == email.lower(),
      Auditor.phone == phone,
    )
  elif email:
    stmt = select(Auditor.id).where(
      Auditor.company_id == company_id,
      Auditor.id != exclude_auditor_id,
      func.lower(Auditor.email) == email.lower(),
    )
  elif phone:
    stmt = select(Auditor.id).where(
      Auditor.company_id == company_id,
      Auditor.id != exclude_auditor_id,
      Auditor.phone == phone,
    )
  else:
    return False

  res = await db.execute(stmt.limit(1))
  return res.scalar_one_or_none() is not None


@router.get(
  '',
  response_model=AuditorsResponse,
  summary='Список аудиторов',
  description='Возвращает аудиторов текущей компании. Поддерживает поиск и фильтры.',
)
async def list_auditors (
  q: str | None = Query(default=None, description='Поиск по ФИО, почте или телефону'),
  city: str | None = Query(default=None, description='Фильтр по городу'),
  gender: AuditorGender | None = Query(default=None, description='Фильтр по полу'),
  birthDateFrom: date | None = Query(default=None, description='Дата рождения от'),
  birthDateTo: date | None = Query(default=None, description='Дата рождения до'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> AuditorsResponse:
  stmt = select(Auditor).where(Auditor.company_id == _company_id(current_user)).order_by(Auditor.created_at.desc())

  if q:
    like = f'%{q.strip()}%'
    stmt = stmt.where(
      or_(
        Auditor.first_name.ilike(like),
        Auditor.last_name.ilike(like),
        func.coalesce(Auditor.middle_name, '').ilike(like),
        func.coalesce(Auditor.email, '').ilike(like),
        func.coalesce(Auditor.phone, '').ilike(like),
      ),
    )
  if city:
    stmt = stmt.where(func.lower(Auditor.city) == city.strip().lower())
  if gender:
    stmt = stmt.where(Auditor.gender == gender)
  if birthDateFrom is not None:
    stmt = stmt.where(Auditor.birth_date >= birthDateFrom)
  if birthDateTo is not None:
    stmt = stmt.where(Auditor.birth_date <= birthDateTo)

  res = await db.execute(stmt)
  return AuditorsResponse(items=[to_auditor_item(a) for a in res.scalars().all()])


@router.get(
  '/export',
  summary='Экспорт аудиторов в Excel (.xlsx)',
)
async def export_auditors (
  q: str | None = Query(default=None, description='Поиск по ФИО, почте или телефону'),
  city: str | None = Query(default=None, description='Фильтр по городу'),
  gender: AuditorGender | None = Query(default=None, description='Фильтр по полу'),
  birthDateFrom: date | None = Query(default=None, description='Дата рождения от'),
  birthDateTo: date | None = Query(default=None, description='Дата рождения до'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> StreamingResponse:
  _ensure_admin(current_user)

  stmt = select(Auditor).where(Auditor.company_id == _company_id(current_user)).order_by(Auditor.created_at.desc())

  if q:
    like = f'%{q.strip()}%'
    stmt = stmt.where(
      or_(
        Auditor.first_name.ilike(like),
        Auditor.last_name.ilike(like),
        func.coalesce(Auditor.middle_name, '').ilike(like),
        func.coalesce(Auditor.email, '').ilike(like),
        func.coalesce(Auditor.phone, '').ilike(like),
      ),
    )
  if city:
    stmt = stmt.where(func.lower(Auditor.city) == city.strip().lower())
  if gender:
    stmt = stmt.where(Auditor.gender == gender)
  if birthDateFrom is not None:
    stmt = stmt.where(Auditor.birth_date >= birthDateFrom)
  if birthDateTo is not None:
    stmt = stmt.where(Auditor.birth_date <= birthDateTo)

  res = await db.execute(stmt)
  items = res.scalars().all()

  headers = ['ID', 'Фамилия', 'Имя', 'Отчество', 'Телефон', 'Email', 'Город', 'Дата рождения', 'Пол', 'Дата создания']
  rows: list[list[object | None]] = []
  for auditor in items:
    rows.append(
      [
        str(auditor.id),
        auditor.last_name,
        auditor.first_name,
        auditor.middle_name,
        auditor.phone,
        auditor.email,
        auditor.city,
        auditor.birth_date.isoformat() if auditor.birth_date else None,
        auditor.gender,
        auditor.created_at.isoformat() if auditor.created_at else None,
      ],
    )

  content = _xlsx_bytes_from_rows(headers, rows)
  filename = f'auditors-{_company_id(current_user)}.xlsx'
  return StreamingResponse(
    io.BytesIO(content),
    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    headers={'Content-Disposition': f'attachment; filename="{filename}"'},
  )


@router.post(
  '',
  response_model=AuditorCreateResponse,
  status_code=status.HTTP_201_CREATED,
  summary='Создать аудитора',
  description='Создаёт аудитора в рамках компании. Доступно только админам.',
)
async def create_auditor (
  payload: AuditorCreateRequest,
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> AuditorCreateResponse:
  _ensure_admin(current_user)

  email = _norm_email(str(payload.email) if payload.email is not None else None)
  phone = _norm_phone(payload.phone)
  if not email and not phone:
    raise HTTPException(status_code=400, detail='Укажите телефон или почту')

  if await _exists_duplicate(db, _company_id(current_user), email, phone):
    raise HTTPException(status_code=400, detail='Аудитор с такими контактами уже существует')

  auditor = Auditor(
    company_id=_company_id(current_user),
    last_name=payload.lastName.strip(),
    first_name=payload.firstName.strip(),
    middle_name=payload.middleName.strip() if payload.middleName else None,
    phone=phone,
    email=email,
    city=payload.city.strip(),
    birth_date=payload.birthDate,
    gender=payload.gender,
  )
  db.add(auditor)
  await db.commit()

  out = await db.execute(select(Auditor).where(Auditor.id == auditor.id))
  auditor = out.scalar_one()
  return AuditorCreateResponse(auditor=to_auditor_item(auditor))


@router.patch(
  '/{auditor_id}',
  response_model=AuditorUpdateResponse,
  summary='Обновить аудитора',
  description='Обновляет аудитора в рамках компании. Доступно только админам.',
)
async def update_auditor (
  payload: AuditorUpdateRequest,
  auditor_id: UUID = Path(..., description='ID аудитора'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> AuditorUpdateResponse:
  _ensure_admin(current_user)

  res = await db.execute(select(Auditor).where(Auditor.id == auditor_id, Auditor.company_id == _company_id(current_user)))
  auditor = res.scalar_one_or_none()
  if auditor is None:
    raise HTTPException(status_code=404, detail='Аудитор не найден')

  email = _norm_email(str(payload.email) if payload.email is not None else None)
  phone = _norm_phone(payload.phone)
  if not email and not phone:
    raise HTTPException(status_code=400, detail='Укажите телефон или почту')

  if await _exists_duplicate_excluding(db, _company_id(current_user), auditor.id, email, phone):
    raise HTTPException(status_code=400, detail='Аудитор с такими контактами уже существует')

  auditor.last_name = payload.lastName.strip()
  auditor.first_name = payload.firstName.strip()
  auditor.middle_name = payload.middleName.strip() if payload.middleName else None
  auditor.phone = phone
  auditor.email = email
  auditor.city = payload.city.strip()
  auditor.birth_date = payload.birthDate
  auditor.gender = payload.gender

  await db.commit()

  out = await db.execute(select(Auditor).where(Auditor.id == auditor.id))
  auditor = out.scalar_one()
  return AuditorUpdateResponse(auditor=to_auditor_item(auditor))


@router.post(
  '/{auditor_id}/password',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Задать пароль аудитору',
  description='Доступно только админам. Нужно для входа аудитора по email+пароль.',
)
async def set_auditor_password (
  payload: AuditorSetPasswordRequest,
  auditor_id: UUID = Path(..., description='ID аудитора'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  _ensure_admin(current_user)

  res = await db.execute(select(Auditor).where(Auditor.id == auditor_id, Auditor.company_id == _company_id(current_user)))
  auditor = res.scalar_one_or_none()
  if auditor is None:
    raise HTTPException(status_code=404, detail='Аудитор не найден')

  auditor.password_hash = hash_password(payload.password)
  await db.commit()
  return None


@router.delete(
  '/{auditor_id}',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Удалить аудитора',
  description='Удаляет аудитора в рамках компании. Доступно только админам.',
)
async def delete_auditor (
  auditor_id: UUID = Path(..., description='ID аудитора'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  _ensure_admin(current_user)

  res = await db.execute(select(Auditor).where(Auditor.id == auditor_id, Auditor.company_id == _company_id(current_user)))
  auditor = res.scalar_one_or_none()
  if auditor is None:
    raise HTTPException(status_code=404, detail='Аудитор не найден')

  await db.delete(auditor)
  await db.commit()
  return None


@router.post(
  '/import',
  response_model=AuditorImportResponse,
  summary='Импорт аудиторов из Excel',
  description='Импортирует аудиторов из .xlsx. Заголовки колонок строго на русском.',
)
async def import_auditors (
  file: UploadFile = File(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> AuditorImportResponse:
  _ensure_admin(current_user)

  filename = (file.filename or '').lower()
  if not filename.endswith('.xlsx'):
    raise HTTPException(status_code=400, detail='Поддерживается только .xlsx')

  raw = await file.read()
  try:
    from openpyxl import load_workbook
  except Exception as e:
    raise HTTPException(status_code=500, detail=f'Не удалось импортировать openpyxl: {e}')

  wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
  ws = wb.active

  expected = ['Фамилия', 'Имя', 'Отчество', 'Телефон', 'Электронная почта', 'Город', 'Дата рождения', 'Пол']
  header: list[str] = []

  created = 0
  skipped = 0
  errors: list[AuditorImportErrorItem] = []

  seen_keys: set[str] = set()

  for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row_idx == 1:
      header = [str(c or '').strip() for c in row]
      if header[: len(expected)] != expected:
        raise HTTPException(status_code=400, detail=f'Неверный заголовок. Ожидается: {", ".join(expected)}')
      continue

    values = list(row)
    rec = {expected[i]: (values[i] if i < len(values) else None) for i in range(len(expected))}

    last_name = str(rec['Фамилия'] or '').strip()
    first_name = str(rec['Имя'] or '').strip()
    middle_name = str(rec['Отчество'] or '').strip() or None
    phone = _norm_phone(str(rec['Телефон'] or '').strip() or None)
    email = _norm_email(str(rec['Электронная почта'] or '').strip() or None)
    city = str(rec['Город'] or '').strip()
    birth_date = _parse_birth_date(rec['Дата рождения'])
    gender = _parse_gender(rec['Пол'])

    # skip empty line
    if not any([last_name, first_name, middle_name, phone, email, city]):
      continue

    if not last_name or not first_name:
      errors.append(AuditorImportErrorItem(row=row_idx, message='Не заполнены «Фамилия»/«Имя»'))
      continue
    if not city:
      errors.append(AuditorImportErrorItem(row=row_idx, message='Не заполнен «Город»'))
      continue
    if birth_date is None:
      errors.append(AuditorImportErrorItem(row=row_idx, message='Не заполнена/не распознана «Дата рождения»'))
      continue
    if gender is None:
      errors.append(AuditorImportErrorItem(row=row_idx, message='Не заполнен/не распознан «Пол» (м/ж)'))
      continue
    if not email and not phone:
      errors.append(AuditorImportErrorItem(row=row_idx, message='Укажите «Телефон» или «Электронная почта»'))
      continue

    key = (email or '') + '|' + (phone or '')
    if key in seen_keys:
      skipped += 1
      continue
    seen_keys.add(key)

    if await _exists_duplicate(db, _company_id(current_user), email, phone):
      skipped += 1
      continue

    auditor = Auditor(
      company_id=_company_id(current_user),
      last_name=last_name,
      first_name=first_name,
      middle_name=middle_name,
      phone=phone,
      email=email,
      city=city,
      birth_date=birth_date,
      gender=gender,
    )
    db.add(auditor)
    created += 1

  await db.commit()
  return AuditorImportResponse(created=created, skipped=skipped, errors=errors)

