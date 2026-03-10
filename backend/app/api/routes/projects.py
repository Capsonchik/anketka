import csv
import hashlib
import io
import os
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Path, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy import delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import get_current_user
from app.db.session import get_db
from app.models.checklist import Checklist
from app.models.checklist_item import ChecklistItem
from app.models.project import Project
from app.models.ref_brand import RefBrand
from app.models.ref_category import RefCategory
from app.models.ref_product import RefProduct
from app.models.ref_city import RefCity
from app.models.ref_region import RefRegion
from app.models.shop_chain import ShopChain
from app.models.shop_point import ShopPoint
from app.models.survey import Survey
from app.models.survey_invite import SurveyInvite
from app.models.project_survey import ProjectSurvey
from app.models.auditor import Auditor
from app.models.project_survey_auditor import ProjectSurveyAuditor
from app.models.user import User
from app.models.user_role import UserRole
from app.schemas.projects import (
  AddressbookCloneRequest,
  AddressbookPointCreateRequest,
  AddressbookPointUpdateRequest,
  AddressbookUploadResponse,
  CreateProjectRequest,
  CreateProjectResponse,
  ChecklistBrandItem,
  ChecklistBrandProducts,
  ChecklistDetailsResponse,
  ChecklistDetailsItem,
  ChecklistItemCreateRequest,
  ChecklistItemPublic,
  ChecklistItemUpdateRequest,
  ChecklistProductItem,
  ChecklistUpdateRequest,
  ChecklistUploadCreatedItem,
  ChecklistUploadResponse,
  ChecklistsResponse,
  ProjectItem,
  ProjectManagerPublic,
  ProjectResponse,
  PointCatalogResponse,
  ProjectsResponse,
  ShopChainItem,
  ShopPointItem,
  ShopPointsResponse,
  RefCitiesResponse,
  RefCityItem,
  RefRegionsResponse,
  RefRegionItem,
)
from app.schemas.auditors import (
  AssignedByItem,
  AuditorItem,
  ProjectSurveyAssignmentItem,
  ProjectSurveyAuditorAssignRequest,
  ProjectSurveyAssignmentsResponse,
)
from app.schemas.surveys import ProjectSurveyAttachRequest, ProjectSurveysResponse, SurveyItem
from app.schemas.surveys import SurveyInviteItem, SurveyInviteResponse


router = APIRouter()


def to_project_item (project: Project) -> ProjectItem:
  manager = project.manager
  manager_public = None
  if manager is not None:
    manager_public = ProjectManagerPublic(
      id=manager.id,
      firstName=manager.first_name,
      lastName=manager.last_name,
      email=manager.email,
    )

  return ProjectItem(
    id=project.id,
    name=project.name,
    comment=project.comment,
    startDate=project.start_date,
    endDate=project.end_date,
    manager=manager_public,
    createdAt=project.created_at,
  )

def to_chain_item (chain: ShopChain) -> ShopChainItem:
  return ShopChainItem(id=chain.id, name=chain.name)

def to_survey_item (s: Survey) -> SurveyItem:
  return SurveyItem(id=s.id, title=s.title, category=s.category, createdAt=s.created_at)

def to_auditor_item (a: Auditor) -> AuditorItem:
  return AuditorItem(
    id=a.id,
    lastName=a.last_name,
    firstName=a.first_name,
    middleName=a.middle_name,
    phone=a.phone,
    email=a.email,
    city=a.city,
    birthDate=a.birth_date,
    gender=(a.gender if a.gender in ('male', 'female') else None),
    createdAt=a.created_at,
  )

def to_assigned_by_item (u: User) -> AssignedByItem:
  return AssignedByItem(id=u.id, firstName=u.first_name, lastName=u.last_name, email=u.email)


def to_shop_point_item (point: ShopPoint) -> ShopPointItem:
  return ShopPointItem(
    id=point.id,
    code=point.code,
    pointName=point.point_name,
    address=point.address,
    regionCode=point.region_code,
    cityName=point.city_name,
    concat=point.concat,
    chain=to_chain_item(point.chain),
  )


def _norm_str (value: object | None) -> str:
  if value is None:
    return ''

  # Excel часто отдаёт целые числа как float (77.0). Для кодов/идентификаторов
  # это критично (например, region_code имеет ограничение по длине).
  if isinstance(value, float) and value.is_integer():
    return str(int(value))

  return str(value).strip()


def _norm_optional (value: object | None) -> str | None:
  v = _norm_str(value)
  return v if v else None


def _lower_key (value: str) -> str:
  return _norm_str(value).lower().replace('ё', 'е')

_REGION_CODE_ALIASES: dict[str, str] = {
  # Moscow and SPB multi-codes (vehicle plates)
  '97': '77',
  '99': '77',
  '177': '77',
  '197': '77',
  '199': '77',
  '777': '77',
  '797': '77',
  '799': '77',
  '977': '77',
  '98': '78',
  '178': '78',
  '198': '78',
  '778': '78',

  # Moscow oblast (vehicle plates)
  '150': '50',
  '190': '50',
  '250': '50',
  '550': '50',
  '750': '50',
  '790': '50',

  # Legacy seeds for new regions
  '180': '93',
  '181': '94',
  '184': '95',
  '185': '90',
  '186': '86',

  # Obsolete autonomous okrugs -> parent regions
  '80': '75',
  '81': '59',
  '82': '41',
  '84': '24',
  '85': '38',
  '88': '24',
}

def _make_autocode (*parts: object | None) -> str | None:
  cleaned: list[str] = []
  for p in parts:
    s = _norm_str(p)
    if s:
      cleaned.append(s)

  if not cleaned:
    return None

  base = '|'.join(cleaned).lower()
  digest = hashlib.blake2s(base.encode('utf-8'), digest_size=6).hexdigest()
  return f'auto-{digest}'


def _guess_ext (filename: str | None) -> str:
  name = (filename or '').lower()
  _, ext = os.path.splitext(name)
  return ext

class _AddressbookRefResolver:
  def __init__ (self, db: AsyncSession):
    self._db = db
    self._city_cache: dict[str, tuple[str | None, str | None]] = {}
    self._region_cache: dict[str, str | None] = {}

  async def _normalize_region_code (self, raw: str) -> str:
    code = raw.strip()
    if not code.isdigit():
      return code

    code = _REGION_CODE_ALIASES.get(code, code)

    # Accept 3-digit plate codes by trying last 2 digits if such subject code exists.
    if len(code) == 3:
      candidate = code[-2:]
      exists = await self._db.execute(select(RefRegion.code).where(RefRegion.code == candidate))
      if exists.scalar_one_or_none() is not None:
        return candidate

    # Accept "1" as "01" (if it exists in regions)
    if len(code) == 1:
      candidate = code.zfill(2)
      exists = await self._db.execute(select(RefRegion.code).where(RefRegion.code == candidate))
      if exists.scalar_one_or_none() is not None:
        return candidate

    return code

  def _key (self, value: str) -> str:
    return _lower_key(value).strip()

  async def resolve (self, city: str | None, region: str | None) -> tuple[str | None, str | None]:
    city_norm = _norm_optional(city)
    region_norm = _norm_optional(region)

    if city_norm:
      key = self._key(city_norm)
      cached = self._city_cache.get(key)
      if cached is not None:
        return cached

      res = await self._db.execute(select(RefCity).where(func.replace(func.lower(RefCity.name), 'ё', 'е') == key))
      city_row = res.scalar_one_or_none()
      if city_row is None:
        region_code, _ = await self.resolve(None, region_norm)
        out = (region_code, None)
        self._city_cache[key] = out
        return out

      out = (city_row.region_code, city_row.name)
      self._city_cache[key] = out
      return out

    if region_norm:
      key = self._key(region_norm)
      cached = self._region_cache.get(key)
      if cached is not None:
        return (cached, None)

      is_code = region_norm.strip().isdigit() and len(region_norm.strip()) <= 3
      if is_code:
        code = await self._normalize_region_code(region_norm)
        stmt = select(RefRegion).where(RefRegion.code == code)
      else:
        stmt = select(RefRegion).where(func.replace(func.lower(RefRegion.name), 'ё', 'е') == key)
      res = await self._db.execute(stmt)
      region_row = res.scalar_one_or_none()
      region_code = region_row.code if region_row is not None else None
      self._region_cache[key] = region_code
      return (region_code, None)

    return (None, None)

async def _make_unique_point_code (db: AsyncSession, project_id: UUID, base_code: str) -> str:
  code = base_code.strip()
  if not code:
    raise HTTPException(status_code=400, detail='Не удалось сформировать код точки')

  res = await db.execute(select(ShopPoint.id).where(ShopPoint.project_id == project_id, ShopPoint.code == code))
  if res.scalar_one_or_none() is None:
    return code

  for i in range(2, 200):
    candidate = f'{code}-{i}'
    r = await db.execute(select(ShopPoint.id).where(ShopPoint.project_id == project_id, ShopPoint.code == candidate))
    if r.scalar_one_or_none() is None:
      return candidate

  raise HTTPException(status_code=400, detail='Не удалось подобрать уникальный код точки')


def _parse_csv_rows (raw: bytes) -> list[dict[str, str]]:
  text = raw.decode('utf-8-sig', errors='ignore')
  reader = csv.DictReader(io.StringIO(text))
  rows: list[dict[str, str]] = []
  for r in reader:
    rows.append({str(k or ''): str(v or '') for k, v in (r or {}).items()})
  return rows


def _parse_xlsx_rows (raw: bytes) -> list[dict[str, str]]:
  try:
    from openpyxl import load_workbook
  except Exception as e:
    raise HTTPException(status_code=500, detail=f'Не удалось импортировать openpyxl: {e}')

  wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
  ws = wb.active
  header: list[str] = []
  rows: list[dict[str, str]] = []

  for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
      header = [str(c or '').strip() for c in row]
      continue
    if not header:
      continue

    rec: dict[str, str] = {}
    has_any = False
    for idx, col in enumerate(header):
      if not col:
        continue
      v = row[idx] if idx < len(row) else None
      s = _norm_str(v)
      if s:
        has_any = True
      rec[col] = s
    if has_any:
      rows.append(rec)

  return rows


def _read_tabular_file (file: UploadFile, raw: bytes) -> list[dict[str, str]]:
  ext = _guess_ext(file.filename)
  if ext in ('.csv',):
    return _parse_csv_rows(raw)
  if ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
    return _parse_xlsx_rows(raw)
  raise HTTPException(status_code=400, detail='Поддерживаются только .xlsx или .csv')


async def _get_project_or_404 (db: AsyncSession, current_user: User, project_id: UUID) -> Project:
  res = await db.execute(
    select(Project)
    .options(selectinload(Project.manager))
    .where(Project.id == project_id, Project.company_id == current_user.company_id),
  )
  project = res.scalar_one_or_none()
  if project is None:
    raise HTTPException(status_code=404, detail='Проект не найден')
  return project


async def _get_or_create_chain (db: AsyncSession, company_id: UUID, name: str) -> ShopChain:
  norm = name.strip()
  res = await db.execute(
    select(ShopChain).where(ShopChain.company_id == company_id, func.lower(ShopChain.name) == norm.lower()),
  )
  chain = res.scalar_one_or_none()
  if chain is not None:
    return chain

  chain = ShopChain(company_id=company_id, name=norm)
  db.add(chain)
  await db.flush()
  return chain


async def _get_or_create_brand (db: AsyncSession, company_id: UUID, name: str) -> RefBrand:
  norm = name.strip()
  res = await db.execute(
    select(RefBrand).where(RefBrand.company_id == company_id, func.lower(RefBrand.name) == norm.lower()),
  )
  brand = res.scalar_one_or_none()
  if brand is not None:
    return brand

  brand = RefBrand(company_id=company_id, name=norm)
  db.add(brand)
  await db.flush()
  return brand


async def _get_or_create_category (db: AsyncSession, company_id: UUID, name: str) -> RefCategory:
  norm = name.strip()
  res = await db.execute(
    select(RefCategory).where(RefCategory.company_id == company_id, func.lower(RefCategory.name) == norm.lower()),
  )
  category = res.scalar_one_or_none()
  if category is not None:
    return category

  category = RefCategory(company_id=company_id, name=norm)
  db.add(category)
  await db.flush()
  return category


async def _get_or_create_product (
  db: AsyncSession,
  brand_id: UUID,
  article: str | None,
  name: str,
  size: str | None,
) -> RefProduct:
  res = await db.execute(
    select(RefProduct).where(
      RefProduct.brand_id == brand_id,
      RefProduct.article.is_(None) if article is None else RefProduct.article == article,
      RefProduct.name == name,
      RefProduct.size.is_(None) if size is None else RefProduct.size == size,
    ),
  )
  product = res.scalar_one_or_none()
  if product is not None:
    return product

  product = RefProduct(brand_id=brand_id, article=article, name=name, size=size)
  db.add(product)
  await db.flush()
  return product


async def _count_checklist_items (db: AsyncSession, checklist_id: UUID) -> int:
  res = await db.execute(select(func.count()).select_from(ChecklistItem).where(ChecklistItem.checklist_id == checklist_id))
  return int(res.scalar_one() or 0)

async def _make_unique_checklist_title (db: AsyncSession, project_id: UUID, chain_id: UUID, base_title: str) -> str:
  title = base_title.strip() or 'checklist'
  suffix = 1
  while True:
    res = await db.execute(
      select(Checklist.id).where(Checklist.project_id == project_id, Checklist.chain_id == chain_id, Checklist.title == title),
    )
    if res.scalar_one_or_none() is None:
      return title
    suffix += 1
    title = f'{base_title}-{suffix}'


@router.get(
  '',
  response_model=ProjectsResponse,
  summary='Список проектов',
  description='Возвращает проекты текущей компании. Поддерживает поиск по названию.',
)
async def list_projects (
  q: str | None = Query(default=None, description='Поиск по названию проекта'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ProjectsResponse:
  stmt = (
    select(Project)
    .options(selectinload(Project.manager))
    .where(Project.company_id == current_user.company_id)
    .order_by(Project.created_at.desc())
  )
  if q:
    like = f'%{q.strip()}%'
    stmt = stmt.where(Project.name.ilike(like))

  res = await db.execute(stmt)
  items = res.scalars().all()
  return ProjectsResponse(items=[to_project_item(p) for p in items])

@router.get(
  '/{project_id}',
  response_model=ProjectResponse,
  summary='Детали проекта',
)
async def get_project (
  project_id: UUID = Path(..., description='ID проекта'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ProjectResponse:
  project = await _get_project_or_404(db, current_user, project_id)
  return ProjectResponse(project=to_project_item(project))

@router.delete(
  '/{project_id}',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Удалить проект',
  description='Удаляет проект текущей компании. Доступно только админам.',
)
async def delete_project (
  project_id: UUID = Path(..., description='ID проекта'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  if current_user.role != UserRole.admin:
    raise HTTPException(status_code=403, detail='Недостаточно прав')

  res = await db.execute(select(Project).where(Project.id == project_id, Project.company_id == current_user.company_id))
  project = res.scalar_one_or_none()
  if project is None:
    raise HTTPException(status_code=404, detail='Проект не найден')

  await db.delete(project)
  await db.commit()
  return None


@router.post(
  '',
  response_model=CreateProjectResponse,
  status_code=status.HTTP_201_CREATED,
  summary='Создать проект',
  description='Создаёт проект в рамках компании текущего пользователя.',
)
async def create_project (
  payload: CreateProjectRequest,
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> CreateProjectResponse:
  if payload.startDate > payload.endDate:
    raise HTTPException(status_code=400, detail='Дата окончания должна быть не раньше даты начала')

  manager_res = await db.execute(
    select(User).where(
      User.id == payload.managerId,
      User.company_id == current_user.company_id,
    ),
  )
  manager = manager_res.scalar_one_or_none()
  if manager is None:
    raise HTTPException(status_code=400, detail='Выбранный менеджер не найден в вашей компании')

  project = Project(
    company_id=current_user.company_id,
    manager_user_id=payload.managerId,
    name=payload.name.strip(),
    comment=payload.comment.strip() if payload.comment else None,
    start_date=payload.startDate,
    end_date=payload.endDate,
  )
  db.add(project)
  await db.commit()

  res = await db.execute(
    select(Project)
    .options(selectinload(Project.manager))
    .where(Project.id == project.id, Project.company_id == current_user.company_id),
  )
  project = res.scalar_one()
  return CreateProjectResponse(project=to_project_item(project))

@router.get(
  '/{project_id}/addressbook',
  response_model=ShopPointsResponse,
  summary='Адресная программа (точки)',
)
async def list_addressbook (
  project_id: UUID = Path(..., description='ID проекта'),
  q: str | None = Query(default=None, description='Поиск по адресу/коду/concat'),
  chainId: UUID | None = Query(default=None, description='Фильтр по сети'),
  regionCode: str | None = Query(default=None, description='Фильтр по коду региона'),
  city: str | None = Query(default=None, description='Фильтр по городу'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ShopPointsResponse:
  await _get_project_or_404(db, current_user, project_id)

  stmt = (
    select(ShopPoint)
    .options(selectinload(ShopPoint.chain))
    .where(ShopPoint.project_id == project_id)
    .order_by(ShopPoint.created_at.desc())
  )
  if chainId is not None:
    stmt = stmt.where(ShopPoint.chain_id == chainId)
  if regionCode:
    stmt = stmt.where(ShopPoint.region_code == regionCode.strip())
  if city:
    stmt = stmt.where(func.lower(ShopPoint.city_name) == city.strip().lower())
  if q:
    like = f'%{q.strip()}%'
    stmt = stmt.where(
      func.coalesce(ShopPoint.address, '').ilike(like)
      | ShopPoint.code.ilike(like)
      | func.coalesce(ShopPoint.concat, '').ilike(like),
    )

  res = await db.execute(stmt)
  items = res.scalars().all()
  return ShopPointsResponse(items=[to_shop_point_item(p) for p in items])

@router.post(
  '/{project_id}/addressbook/points',
  response_model=ShopPointItem,
  status_code=status.HTTP_201_CREATED,
  summary='Добавить точку в адресную программу',
)
async def create_addressbook_point (
  payload: AddressbookPointCreateRequest,
  project_id: UUID = Path(..., description='ID проекта'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ShopPointItem:
  await _get_project_or_404(db, current_user, project_id)

  resolver = _AddressbookRefResolver(db)
  region_code, city_name = await resolver.resolve(payload.city, payload.region)

  chain = await _get_or_create_chain(db, current_user.company_id, payload.chainName)

  desired = _norm_optional(payload.code)
  if desired:
    # enforce uniqueness within project
    exists = await db.execute(select(ShopPoint.id).where(ShopPoint.project_id == project_id, ShopPoint.code == desired))
    if exists.scalar_one_or_none() is not None:
      raise HTTPException(status_code=400, detail='Код точки уже используется в этом проекте')
    code = desired
  else:
    base_code = _make_autocode(payload.chainName, region_code or payload.region, city_name or payload.city, payload.address, payload.concat)
    code = await _make_unique_point_code(db, project_id, base_code or '')

  point = ShopPoint(
    project_id=project_id,
    chain_id=chain.id,
    code=code,
    point_name=None,
    address=payload.address.strip(),
    region_code=region_code,
    city_name=city_name,
    concat=payload.concat.strip() if payload.concat else None,
  )
  db.add(point)
  await db.commit()

  res = await db.execute(select(ShopPoint).options(selectinload(ShopPoint.chain)).where(ShopPoint.id == point.id))
  point = res.scalar_one()
  return to_shop_point_item(point)

@router.patch(
  '/{project_id}/addressbook/points/{point_id}',
  response_model=ShopPointItem,
  summary='Обновить точку адресной программы',
)
async def update_addressbook_point (
  payload: AddressbookPointUpdateRequest,
  project_id: UUID = Path(..., description='ID проекта'),
  point_id: UUID = Path(..., description='ID точки'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ShopPointItem:
  await _get_project_or_404(db, current_user, project_id)

  res = await db.execute(
    select(ShopPoint)
    .options(selectinload(ShopPoint.chain))
    .where(ShopPoint.id == point_id, ShopPoint.project_id == project_id),
  )
  point = res.scalar_one_or_none()
  if point is None:
    raise HTTPException(status_code=404, detail='Точка не найдена')

  resolver = _AddressbookRefResolver(db)
  region_code, city_name = await resolver.resolve(payload.city, payload.region)
  chain = await _get_or_create_chain(db, current_user.company_id, payload.chainName)

  desired = _norm_optional(payload.code)
  if desired and desired != point.code:
    exists = await db.execute(
      select(ShopPoint.id).where(
        ShopPoint.project_id == project_id,
        ShopPoint.code == desired,
        ShopPoint.id != point.id,
      ),
    )
    if exists.scalar_one_or_none() is not None:
      raise HTTPException(status_code=400, detail='Код точки уже используется в этом проекте')
    point.code = desired

  point.chain_id = chain.id
  point.address = payload.address.strip()
  point.region_code = region_code
  point.city_name = city_name
  point.concat = payload.concat.strip() if payload.concat else None

  await db.commit()

  out_res = await db.execute(select(ShopPoint).options(selectinload(ShopPoint.chain)).where(ShopPoint.id == point.id))
  point = out_res.scalar_one()
  return to_shop_point_item(point)

@router.delete(
  '/{project_id}/addressbook/points/{point_id}',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Удалить точку адресной программы',
)
async def delete_addressbook_point (
  project_id: UUID = Path(..., description='ID проекта'),
  point_id: UUID = Path(..., description='ID точки'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  await _get_project_or_404(db, current_user, project_id)

  res = await db.execute(select(ShopPoint).where(ShopPoint.id == point_id, ShopPoint.project_id == project_id))
  point = res.scalar_one_or_none()
  if point is None:
    raise HTTPException(status_code=404, detail='Точка не найдена')

  await db.delete(point)
  await db.commit()
  return None


@router.post(
  '/{project_id}/addressbook/upload',
  response_model=AddressbookUploadResponse,
  summary='Загрузить адресный справочник (файл)',
  description='Принимает Excel/CSV файл и загружает точки проекта. Сеть/магазин связывается через ShopChain.',
)
async def upload_addressbook (
  project_id: UUID = Path(..., description='ID проекта'),
  file: UploadFile = File(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> AddressbookUploadResponse:
  await _get_project_or_404(db, current_user, project_id)

  raw = await file.read()
  rows = _read_tabular_file(file, raw)
  if not rows:
    raise HTTPException(status_code=400, detail='Файл пустой или не удалось прочитать строки')

  # Маппинг колонок (варианты названий)
  def pick (row: dict[str, str], keys: list[str]) -> str | None:
    for k in keys:
      for rk, rv in row.items():
        if _lower_key(rk) == _lower_key(k):
          v = _norm_optional(rv)
          if v:
            return v
    return None

  resolver = _AddressbookRefResolver(db)

  created = 0
  updated = 0

  for r in rows:
    chain_name = pick(r, ['магазин', 'сеть', 'shop', 'shop_chain', 'name'])
    code = pick(r, ['code', 'код', 'id', 'точка', 'point', 'point_code'])
    address = pick(r, ['address', 'адрес'])
    region_in = pick(r, ['region', 'регион', 'region_code', 'reg_id'])
    city_in = pick(r, ['city', 'город'])
    concat = pick(r, ['concat'])

    if not chain_name:
      continue

    region_code, city_name = await resolver.resolve(city_in, region_in)

    if not code:
      code = _make_autocode(chain_name, region_code or region_in, city_name or city_in, address, concat)
      if not code:
        continue

    chain = await _get_or_create_chain(db, current_user.company_id, chain_name)

    existing_res = await db.execute(
      select(ShopPoint).where(ShopPoint.project_id == project_id, ShopPoint.code == code),
    )
    point = existing_res.scalar_one_or_none()
    if point is None:
      point = ShopPoint(
        project_id=project_id,
        chain_id=chain.id,
        code=code,
        point_name=None,
        address=address,
        region_code=region_code,
        city_name=city_name,
        concat=concat,
      )
      db.add(point)
      created += 1
    else:
      point.chain_id = chain.id
      point.address = address
      point.region_code = region_code
      point.city_name = city_name
      point.concat = concat
      updated += 1

  await db.commit()
  return AddressbookUploadResponse(created=created, updated=updated)


@router.post(
  '/{project_id}/addressbook/clone',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Скопировать адресный справочник из проекта',
  description='Прототип. Проверяет доступ к проектам и готовит дальнейшее копирование.',
)
async def clone_addressbook (
  payload: AddressbookCloneRequest,
  project_id: UUID = Path(..., description='ID проекта'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  res = await db.execute(select(Project).where(Project.id == project_id, Project.company_id == current_user.company_id))
  target = res.scalar_one_or_none()
  if target is None:
    raise HTTPException(status_code=404, detail='Проект не найден')

  src_res = await db.execute(
    select(Project).where(Project.id == payload.sourceProjectId, Project.company_id == current_user.company_id),
  )
  source = src_res.scalar_one_or_none()
  if source is None:
    raise HTTPException(status_code=400, detail='Проект-источник не найден')

  return None


@router.get(
  '/{project_id}/checklists',
  response_model=ChecklistsResponse,
  summary='Список чек-листов проекта',
)
async def list_checklists (
  project_id: UUID = Path(..., description='ID проекта'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ChecklistsResponse:
  await _get_project_or_404(db, current_user, project_id)
  res = await db.execute(
    select(Checklist)
    .options(selectinload(Checklist.chain))
    .where(Checklist.project_id == project_id)
    .order_by(Checklist.created_at.desc()),
  )
  items = res.scalars().all()

  out: list[ChecklistItemPublic] = []
  for cl in items:
    out.append(
      ChecklistItemPublic(
        id=cl.id,
        title=cl.title,
        chain=to_chain_item(cl.chain),
        itemsCount=await _count_checklist_items(db, cl.id),
        createdAt=cl.created_at,
      ),
    )

  return ChecklistsResponse(items=out)


def _xlsx_bytes_from_rows (headers: list[str], rows: list[list[object | None]]) -> bytes:
  try:
    from openpyxl import Workbook
  except Exception as e:
    raise HTTPException(status_code=500, detail=f'Не удалось импортировать openpyxl: {e}')

  wb = Workbook(write_only=True)
  ws = wb.create_sheet(title='Sheet1')
  ws.append(headers)
  for r in rows:
    ws.append([('' if v is None else v) for v in r])

  buf = io.BytesIO()
  wb.save(buf)
  return buf.getvalue()


@router.get(
  '/{project_id}/addressbook/export',
  summary='Экспорт адресной программы (xlsx)',
)
async def export_addressbook_xlsx (
  project_id: UUID = Path(..., description='ID проекта'),
  q: str | None = Query(default=None, description='Поиск по адресу/коду/concat'),
  chainId: UUID | None = Query(default=None, description='Фильтр по сети'),
  regionCode: str | None = Query(default=None, description='Фильтр по коду региона'),
  city: str | None = Query(default=None, description='Фильтр по городу'),
  format: str | None = Query(default=None, description='Формат (xlsx)'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> StreamingResponse:
  await _get_project_or_404(db, current_user, project_id)

  stmt = (
    select(ShopPoint)
    .options(selectinload(ShopPoint.chain))
    .where(ShopPoint.project_id == project_id)
    .order_by(ShopPoint.created_at.desc())
  )
  if chainId is not None:
    stmt = stmt.where(ShopPoint.chain_id == chainId)
  if regionCode:
    stmt = stmt.where(ShopPoint.region_code == regionCode.strip())
  if city:
    stmt = stmt.where(func.lower(ShopPoint.city_name) == city.strip().lower())
  if q:
    like = f'%{q.strip()}%'
    stmt = stmt.where(
      func.coalesce(ShopPoint.address, '').ilike(like)
      | ShopPoint.code.ilike(like)
      | func.coalesce(ShopPoint.concat, '').ilike(like),
    )

  res = await db.execute(stmt)
  items = res.scalars().all()

  headers = ['code', 'chain', 'city', 'region', 'address', 'concat']
  rows: list[list[object | None]] = []
  for p in items:
    rows.append([p.code, p.chain.name if p.chain else None, p.city_name, p.region_code, p.address, p.concat])

  content = _xlsx_bytes_from_rows(headers, rows)
  filename = f'addressbook-{project_id}.xlsx'
  return StreamingResponse(
    io.BytesIO(content),
    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    headers={'Content-Disposition': f'attachment; filename="{filename}"'},
  )


@router.get(
  '/{project_id}/checklists/export',
  summary='Экспорт чек-листов (xlsx)',
)
async def export_checklists_xlsx (
  project_id: UUID = Path(..., description='ID проекта'),
  q: str | None = Query(default=None, description='Поиск по названию/сети'),
  chainId: UUID | None = Query(default=None, description='Фильтр по сети'),
  format: str | None = Query(default=None, description='Формат (xlsx)'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> StreamingResponse:
  await _get_project_or_404(db, current_user, project_id)

  items_count = func.count(ChecklistItem.id).label('items_count')
  stmt = (
    select(Checklist, ShopChain, items_count)
    .join(ShopChain, ShopChain.id == Checklist.chain_id)
    .outerjoin(ChecklistItem, ChecklistItem.checklist_id == Checklist.id)
    .where(Checklist.project_id == project_id)
    .group_by(Checklist.id, ShopChain.id)
    .order_by(Checklist.created_at.desc())
  )
  if chainId is not None:
    stmt = stmt.where(Checklist.chain_id == chainId)
  if q:
    like = f'%{q.strip()}%'
    stmt = stmt.where(Checklist.title.ilike(like) | ShopChain.name.ilike(like))

  res = await db.execute(stmt)
  rows_db = res.all()

  headers = ['chain', 'title', 'itemsCount', 'createdAt']
  rows: list[list[object | None]] = []
  for checklist, chain, cnt in rows_db:
    rows.append([chain.name if chain else None, checklist.title, int(cnt or 0), checklist.created_at.isoformat()])

  content = _xlsx_bytes_from_rows(headers, rows)
  filename = f'checklists-{project_id}.xlsx'
  return StreamingResponse(
    io.BytesIO(content),
    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    headers={'Content-Disposition': f'attachment; filename="{filename}"'},
  )


@router.get(
  '/{project_id}/checklists/{checklist_id}/items/export',
  summary='Экспорт позиций чек-листа (xlsx)',
)
async def export_checklist_items_xlsx (
  project_id: UUID = Path(..., description='ID проекта'),
  checklist_id: UUID = Path(..., description='ID чек-листа'),
  q: str | None = Query(default=None, description='Поиск по категории/бренду/артикулу/названию/размеру'),
  brand: str | None = Query(default=None, description='Фильтр по бренду'),
  category: str | None = Query(default=None, description='Фильтр по категории'),
  format: str | None = Query(default=None, description='Формат (xlsx)'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> StreamingResponse:
  await _get_project_or_404(db, current_user, project_id)

  cl_res = await db.execute(select(Checklist.id).where(Checklist.id == checklist_id, Checklist.project_id == project_id))
  if cl_res.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Чек-лист не найден')

  stmt = (
    select(ChecklistItem, RefCategory, RefProduct, RefBrand)
    .join(RefCategory, RefCategory.id == ChecklistItem.category_id)
    .join(RefProduct, RefProduct.id == ChecklistItem.product_id)
    .join(RefBrand, RefBrand.id == RefProduct.brand_id)
    .where(ChecklistItem.checklist_id == checklist_id)
    .order_by(ChecklistItem.sort_order.asc())
  )
  if brand:
    stmt = stmt.where(RefBrand.name == brand)
  if category:
    stmt = stmt.where(RefCategory.name == category)
  if q:
    like = f'%{q.strip()}%'
    stmt = stmt.where(
      RefCategory.name.ilike(like)
      | RefBrand.name.ilike(like)
      | func.coalesce(RefProduct.article, '').ilike(like)
      | RefProduct.name.ilike(like)
      | func.coalesce(RefProduct.size, '').ilike(like),
    )

  res = await db.execute(stmt)
  rows_db = res.all()

  headers = ['category', 'brand', 'article', 'name', 'size']
  rows: list[list[object | None]] = []
  for _, cat, prod, br in rows_db:
    rows.append([cat.name, br.name, prod.article, prod.name, prod.size])

  content = _xlsx_bytes_from_rows(headers, rows)
  filename = f'checklist-items-{checklist_id}.xlsx'
  return StreamingResponse(
    io.BytesIO(content),
    media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    headers={'Content-Disposition': f'attachment; filename="{filename}"'},
  )


@router.get(
  '/{project_id}/surveys',
  response_model=ProjectSurveysResponse,
  summary='Анкеты проекта',
)
async def list_project_surveys (
  project_id: UUID = Path(..., description='ID проекта'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ProjectSurveysResponse:
  await _get_project_or_404(db, current_user, project_id)
  res = await db.execute(
    select(Survey)
    .join(ProjectSurvey, ProjectSurvey.survey_id == Survey.id)
    .where(ProjectSurvey.project_id == project_id, Survey.company_id == current_user.company_id)
    .order_by(ProjectSurvey.attached_at.desc()),
  )
  return ProjectSurveysResponse(items=[to_survey_item(x) for x in res.scalars().all()])


@router.post(
  '/{project_id}/surveys',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Прикрепить анкету к проекту',
)
async def attach_project_survey (
  payload: ProjectSurveyAttachRequest,
  project_id: UUID = Path(..., description='ID проекта'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  await _get_project_or_404(db, current_user, project_id)

  sres = await db.execute(select(Survey).where(Survey.id == payload.surveyId, Survey.company_id == current_user.company_id))
  survey = sres.scalar_one_or_none()
  if survey is None:
    raise HTTPException(status_code=404, detail='Анкета не найдена')

  exists = await db.execute(
    select(ProjectSurvey.project_id).where(ProjectSurvey.project_id == project_id, ProjectSurvey.survey_id == survey.id),
  )
  if exists.scalar_one_or_none() is not None:
    return None

  link = ProjectSurvey(project_id=project_id, survey_id=survey.id)
  db.add(link)
  await db.commit()
  return None


@router.delete(
  '/{project_id}/surveys/{survey_id}',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Открепить анкету от проекта',
)
async def detach_project_survey (
  project_id: UUID = Path(..., description='ID проекта'),
  survey_id: UUID = Path(..., description='ID анкеты'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  await _get_project_or_404(db, current_user, project_id)

  # Проверим, что анкета принадлежит компании (чтобы не отдавать existence по чужим id)
  sres = await db.execute(select(Survey.id).where(Survey.id == survey_id, Survey.company_id == current_user.company_id))
  if sres.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Анкета не найдена')

  res = await db.execute(
    select(ProjectSurvey).where(ProjectSurvey.project_id == project_id, ProjectSurvey.survey_id == survey_id),
  )
  link = res.scalar_one_or_none()
  if link is None:
    raise HTTPException(status_code=404, detail='Анкета не прикреплена к проекту')

  await db.delete(link)
  await db.commit()
  return None


@router.get(
  '/{project_id}/surveys/{survey_id}/auditors',
  response_model=ProjectSurveyAssignmentsResponse,
  summary='Назначенные аудиторы (проект+анкета)',
)
async def list_project_survey_auditors (
  project_id: UUID = Path(..., description='ID проекта'),
  survey_id: UUID = Path(..., description='ID анкеты'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ProjectSurveyAssignmentsResponse:
  await _get_project_or_404(db, current_user, project_id)

  # анкета должна быть прикреплена к проекту
  link_res = await db.execute(select(ProjectSurvey.project_id).where(ProjectSurvey.project_id == project_id, ProjectSurvey.survey_id == survey_id))
  if link_res.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Анкета не прикреплена к проекту')

  res = await db.execute(
    select(ProjectSurveyAuditor, Auditor, User)
    .join(Auditor, Auditor.id == ProjectSurveyAuditor.auditor_id)
    .outerjoin(User, User.id == ProjectSurveyAuditor.assigned_by_user_id)
    .where(
      ProjectSurveyAuditor.project_id == project_id,
      ProjectSurveyAuditor.survey_id == survey_id,
      Auditor.company_id == current_user.company_id,
    )
    .order_by(ProjectSurveyAuditor.assigned_at.desc(), Auditor.last_name.asc(), Auditor.first_name.asc()),
  )

  items: list[ProjectSurveyAssignmentItem] = []
  for link, auditor, assigned_by in res.all():
    items.append(
      ProjectSurveyAssignmentItem(
        auditor=to_auditor_item(auditor),
        assignedAt=link.assigned_at,
        assignedBy=to_assigned_by_item(assigned_by) if assigned_by is not None else None,
        status=link.status,
        acceptedAt=link.accepted_at,
        declinedAt=link.declined_at,
        inProgressAt=link.in_progress_at,
      ),
    )

  return ProjectSurveyAssignmentsResponse(items=items)


@router.post(
  '/{project_id}/surveys/{survey_id}/auditors',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Назначить аудиторов (проект+анкета)',
  description='Идемпотентно назначает аудиторов на связку проект+анкета. Доступно только админам.',
)
async def assign_project_survey_auditors (
  payload: ProjectSurveyAuditorAssignRequest,
  project_id: UUID = Path(..., description='ID проекта'),
  survey_id: UUID = Path(..., description='ID анкеты'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  if current_user.role != UserRole.admin:
    raise HTTPException(status_code=403, detail='Недостаточно прав')

  await _get_project_or_404(db, current_user, project_id)

  link_res = await db.execute(select(ProjectSurvey.project_id).where(ProjectSurvey.project_id == project_id, ProjectSurvey.survey_id == survey_id))
  if link_res.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Анкета не прикреплена к проекту')

  auditor_ids = list(dict.fromkeys(payload.auditorIds))
  if not auditor_ids:
    return None

  # только аудиторы своей компании
  existing_auditors = await db.execute(select(Auditor.id).where(Auditor.company_id == current_user.company_id, Auditor.id.in_(auditor_ids)))
  allowed_ids = set(existing_auditors.scalars().all())
  if not allowed_ids:
    raise HTTPException(status_code=400, detail='Аудиторы не найдены')

  # уже назначенные
  existing_links = await db.execute(
    select(ProjectSurveyAuditor.auditor_id).where(
      ProjectSurveyAuditor.project_id == project_id,
      ProjectSurveyAuditor.survey_id == survey_id,
      ProjectSurveyAuditor.auditor_id.in_(list(allowed_ids)),
    ),
  )
  assigned = set(existing_links.scalars().all())

  for aid in allowed_ids:
    if aid in assigned:
      continue
    db.add(ProjectSurveyAuditor(project_id=project_id, survey_id=survey_id, auditor_id=aid, assigned_by_user_id=current_user.id))

  await db.commit()
  return None


@router.delete(
  '/{project_id}/surveys/{survey_id}/auditors/{auditor_id}',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Снять назначение аудитора',
  description='Снимает назначение аудитора с проект+анкета. Доступно только админам.',
)
async def unassign_project_survey_auditor (
  project_id: UUID = Path(..., description='ID проекта'),
  survey_id: UUID = Path(..., description='ID анкеты'),
  auditor_id: UUID = Path(..., description='ID аудитора'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  if current_user.role != UserRole.admin:
    raise HTTPException(status_code=403, detail='Недостаточно прав')

  await _get_project_or_404(db, current_user, project_id)

  res = await db.execute(
    select(ProjectSurveyAuditor).where(
      ProjectSurveyAuditor.project_id == project_id,
      ProjectSurveyAuditor.survey_id == survey_id,
      ProjectSurveyAuditor.auditor_id == auditor_id,
    ),
  )
  link = res.scalar_one_or_none()
  if link is None:
    raise HTTPException(status_code=404, detail='Назначение не найдено')

  await db.delete(link)
  await db.commit()
  return None


@router.delete(
  '/{project_id}/surveys/{survey_id}/auditors',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Снять все назначения аудиторов',
  description='Снимает все назначения аудиторов с проект+анкета. Доступно только админам.',
)
async def unassign_all_project_survey_auditors (
  project_id: UUID = Path(..., description='ID проекта'),
  survey_id: UUID = Path(..., description='ID анкеты'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  if current_user.role != UserRole.admin:
    raise HTTPException(status_code=403, detail='Недостаточно прав')

  await _get_project_or_404(db, current_user, project_id)

  await db.execute(
    delete(ProjectSurveyAuditor).where(
      ProjectSurveyAuditor.project_id == project_id,
      ProjectSurveyAuditor.survey_id == survey_id,
    ),
  )
  await db.commit()
  return None


def _ensure_manager_or_admin (user: User) -> None:
  if user.role not in (UserRole.admin, UserRole.manager):
    raise HTTPException(status_code=403, detail='Недостаточно прав')


async def _get_or_create_invite (
  db: AsyncSession,
  *,
  project_id: UUID,
  survey_id: UUID,
  auditor_id: UUID | None,
  purpose: str,
  created_by_user_id: UUID | None,
) -> SurveyInvite:
  res = await db.execute(
    select(SurveyInvite)
    .where(
      SurveyInvite.project_id == project_id,
      SurveyInvite.survey_id == survey_id,
      SurveyInvite.auditor_id.is_(None) if auditor_id is None else SurveyInvite.auditor_id == auditor_id,
      SurveyInvite.purpose == purpose,
      SurveyInvite.revoked_at.is_(None),
    )
    .order_by(SurveyInvite.created_at.desc())
    .limit(1),
  )
  existing = res.scalar_one_or_none()
  if existing is not None:
    return existing

  invite = SurveyInvite(
    project_id=project_id,
    survey_id=survey_id,
    auditor_id=auditor_id,
    purpose=purpose,
    created_by_user_id=created_by_user_id,
  )
  db.add(invite)
  await db.commit()
  out = await db.execute(select(SurveyInvite).where(SurveyInvite.token == invite.token))
  return out.scalar_one()


def to_invite_item (inv: SurveyInvite) -> SurveyInviteItem:
  return SurveyInviteItem(
    token=inv.token,
    projectId=inv.project_id,
    surveyId=inv.survey_id,
    auditorId=inv.auditor_id,
    purpose=inv.purpose,
    createdAt=inv.created_at,
  )


@router.post(
  '/{project_id}/surveys/{survey_id}/invites/test',
  response_model=SurveyInviteResponse,
  summary='Тестовая ссылка на прохождение (проект+анкета)',
  description='Создаёт/возвращает тестовую ссылку (invite token) для прохождения анкеты в контексте проекта.',
)
async def create_test_invite (
  project_id: UUID = Path(..., description='ID проекта'),
  survey_id: UUID = Path(..., description='ID анкеты'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> SurveyInviteResponse:
  _ensure_manager_or_admin(current_user)
  await _get_project_or_404(db, current_user, project_id)

  link_res = await db.execute(select(ProjectSurvey.project_id).where(ProjectSurvey.project_id == project_id, ProjectSurvey.survey_id == survey_id))
  if link_res.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Анкета не прикреплена к проекту')

  invite = await _get_or_create_invite(
    db,
    project_id=project_id,
    survey_id=survey_id,
    auditor_id=None,
    purpose='test',
    created_by_user_id=current_user.id,
  )
  return SurveyInviteResponse(invite=to_invite_item(invite))


@router.post(
  '/{project_id}/surveys/{survey_id}/auditors/{auditor_id}/invite',
  response_model=SurveyInviteResponse,
  summary='Ссылка для аудитора (проект+анкета)',
  description='Создаёт/возвращает ссылку на прохождение анкеты для назначенного аудитора.',
)
async def create_auditor_invite (
  project_id: UUID = Path(..., description='ID проекта'),
  survey_id: UUID = Path(..., description='ID анкеты'),
  auditor_id: UUID = Path(..., description='ID аудитора'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> SurveyInviteResponse:
  _ensure_manager_or_admin(current_user)
  await _get_project_or_404(db, current_user, project_id)

  link_res = await db.execute(select(ProjectSurvey.project_id).where(ProjectSurvey.project_id == project_id, ProjectSurvey.survey_id == survey_id))
  if link_res.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Анкета не прикреплена к проекту')

  assigned_res = await db.execute(
    select(ProjectSurveyAuditor.auditor_id).where(
      ProjectSurveyAuditor.project_id == project_id,
      ProjectSurveyAuditor.survey_id == survey_id,
      ProjectSurveyAuditor.auditor_id == auditor_id,
    ),
  )
  if assigned_res.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Аудитор не назначен на эту анкету')

  # аудитора проверяем, чтобы не светить existence чужих id
  aud_res = await db.execute(select(Auditor.id).where(Auditor.id == auditor_id, Auditor.company_id == current_user.company_id))
  if aud_res.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Аудитор не найден')

  invite = await _get_or_create_invite(
    db,
    project_id=project_id,
    survey_id=survey_id,
    auditor_id=auditor_id,
    purpose='assignment',
    created_by_user_id=current_user.id,
  )
  return SurveyInviteResponse(invite=to_invite_item(invite))


@router.post(
  '/{project_id}/checklists/upload',
  response_model=ChecklistUploadResponse,
  summary='Загрузить чек-лист (файл)',
  description='Загружает чек-лист(ы) проекта. Если в файле несколько магазинов/сетей — создаст несколько чек-листов (по сети).',
)
async def upload_checklist (
  project_id: UUID = Path(..., description='ID проекта'),
  file: UploadFile = File(...),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ChecklistUploadResponse:
  await _get_project_or_404(db, current_user, project_id)

  raw = await file.read()
  rows = _read_tabular_file(file, raw)
  if not rows:
    raise HTTPException(status_code=400, detail='Файл пустой или не удалось прочитать строки')

  # Нормализуем заголовки на рус/англ
  def val (row: dict[str, str], keys: list[str]) -> str:
    for k in keys:
      for rk, rv in row.items():
        if _lower_key(rk) == _lower_key(k):
          return _norm_str(rv)
    return ''

  # группировка по сети
  grouped: dict[str, list[dict[str, str]]] = {}
  for r in rows:
    store = val(r, ['магазин', 'store', 'shop'])
    if not store:
      continue
    grouped.setdefault(store.strip(), []).append(r)

  if not grouped:
    raise HTTPException(status_code=400, detail='Не найдена колонка \"Магазин\"/\"store\" или строки пустые')

  created: list[ChecklistUploadCreatedItem] = []
  base_title = os.path.splitext(os.path.basename(file.filename or 'checklist'))[0] or 'checklist'

  for store_name, store_rows in grouped.items():
    chain = await _get_or_create_chain(db, current_user.company_id, store_name)
    title = await _make_unique_checklist_title(db, project_id, chain.id, base_title)

    checklist = Checklist(project_id=project_id, chain_id=chain.id, title=title)
    db.add(checklist)
    await db.flush()

    sort_order = 0
    for r in store_rows:
      category_name = val(r, ['категория', 'category'])
      brand_name = val(r, ['брэнд', 'бренд', 'brand'])
      article = _norm_optional(val(r, ['артикул', 'sku', 'article']))
      product_name = val(r, ['название', 'name'])
      size = _norm_optional(val(r, ['размер', 'size']))

      if not category_name or not brand_name or not product_name:
        continue

      category = await _get_or_create_category(db, current_user.company_id, category_name)
      brand = await _get_or_create_brand(db, current_user.company_id, brand_name)
      product = await _get_or_create_product(db, brand.id, article, product_name.strip(), size)

      item = ChecklistItem(checklist_id=checklist.id, category_id=category.id, product_id=product.id, sort_order=sort_order)
      db.add(item)
      sort_order += 1

    await db.flush()
    created.append(
      ChecklistUploadCreatedItem(
        checklistId=checklist.id,
        chain=to_chain_item(chain),
        title=title,
        itemsCount=await _count_checklist_items(db, checklist.id),
      ),
    )

  await db.commit()
  return ChecklistUploadResponse(created=created)


@router.get(
  '/{project_id}/checklists/{checklist_id}',
  response_model=ChecklistDetailsResponse,
  summary='Детали чек-листа',
)
async def get_checklist (
  project_id: UUID = Path(..., description='ID проекта'),
  checklist_id: UUID = Path(..., description='ID чек-листа'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ChecklistDetailsResponse:
  await _get_project_or_404(db, current_user, project_id)
  res = await db.execute(
    select(Checklist)
    .options(selectinload(Checklist.chain))
    .where(Checklist.id == checklist_id, Checklist.project_id == project_id),
  )
  checklist = res.scalar_one_or_none()
  if checklist is None:
    raise HTTPException(status_code=404, detail='Чек-лист не найден')

  items_res = await db.execute(
    select(ChecklistItem, RefCategory, RefProduct, RefBrand)
    .join(RefCategory, RefCategory.id == ChecklistItem.category_id)
    .join(RefProduct, RefProduct.id == ChecklistItem.product_id)
    .join(RefBrand, RefBrand.id == RefProduct.brand_id)
    .where(ChecklistItem.checklist_id == checklist.id)
    .order_by(ChecklistItem.sort_order.asc()),
  )

  items: list[ChecklistDetailsItem] = []
  for item, category, product, brand in items_res.all():
    items.append(
      ChecklistDetailsItem(
        itemId=item.id,
        sortOrder=item.sort_order,
        category=category.name,
        brand=brand.name,
        article=product.article,
        name=product.name,
        size=product.size,
      ),
    )

  return ChecklistDetailsResponse(
    checklistId=checklist.id,
    title=checklist.title,
    chain=to_chain_item(checklist.chain),
    itemsCount=await _count_checklist_items(db, checklist.id),
    createdAt=checklist.created_at,
    items=items,
  )

@router.post(
  '/{project_id}/checklists/{checklist_id}/items',
  response_model=ChecklistDetailsItem,
  status_code=status.HTTP_201_CREATED,
  summary='Добавить строку в чек-лист',
)
async def create_checklist_item (
  payload: ChecklistItemCreateRequest,
  project_id: UUID = Path(..., description='ID проекта'),
  checklist_id: UUID = Path(..., description='ID чек-листа'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ChecklistDetailsItem:
  await _get_project_or_404(db, current_user, project_id)

  cl_res = await db.execute(select(Checklist.id).where(Checklist.id == checklist_id, Checklist.project_id == project_id))
  if cl_res.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Чек-лист не найден')

  category = await _get_or_create_category(db, current_user.company_id, payload.category)
  brand = await _get_or_create_brand(db, current_user.company_id, payload.brand)
  product = await _get_or_create_product(
    db,
    brand.id,
    _norm_optional(payload.article),
    payload.name.strip(),
    _norm_optional(payload.size),
  )

  # не дублируем позицию
  exists_res = await db.execute(
    select(ChecklistItem.id).where(
      ChecklistItem.checklist_id == checklist_id,
      ChecklistItem.category_id == category.id,
      ChecklistItem.product_id == product.id,
    ),
  )
  if exists_res.scalar_one_or_none() is not None:
    raise HTTPException(status_code=400, detail='Такая позиция уже есть в чек-листе')

  sort_order = payload.sortOrder
  if sort_order is None:
    max_res = await db.execute(select(func.max(ChecklistItem.sort_order)).where(ChecklistItem.checklist_id == checklist_id))
    sort_order = int((max_res.scalar_one_or_none() or 0)) + 1

  item = ChecklistItem(checklist_id=checklist_id, category_id=category.id, product_id=product.id, sort_order=int(sort_order))
  db.add(item)
  await db.commit()

  out_res = await db.execute(
    select(ChecklistItem, RefCategory, RefProduct, RefBrand)
    .join(RefCategory, RefCategory.id == ChecklistItem.category_id)
    .join(RefProduct, RefProduct.id == ChecklistItem.product_id)
    .join(RefBrand, RefBrand.id == RefProduct.brand_id)
    .where(ChecklistItem.id == item.id),
  )
  out_item, out_category, out_product, out_brand = out_res.one()
  return ChecklistDetailsItem(
    itemId=out_item.id,
    sortOrder=out_item.sort_order,
    category=out_category.name,
    brand=out_brand.name,
    article=out_product.article,
    name=out_product.name,
    size=out_product.size,
  )


@router.patch(
  '/{project_id}/checklists/{checklist_id}/items/{item_id}',
  response_model=ChecklistDetailsItem,
  summary='Обновить строку чек-листа',
)
async def update_checklist_item (
  payload: ChecklistItemUpdateRequest,
  project_id: UUID = Path(..., description='ID проекта'),
  checklist_id: UUID = Path(..., description='ID чек-листа'),
  item_id: UUID = Path(..., description='ID строки'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ChecklistDetailsItem:
  await _get_project_or_404(db, current_user, project_id)

  cl_res = await db.execute(select(Checklist.id).where(Checklist.id == checklist_id, Checklist.project_id == project_id))
  if cl_res.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Чек-лист не найден')

  item_res = await db.execute(select(ChecklistItem).where(ChecklistItem.id == item_id, ChecklistItem.checklist_id == checklist_id))
  item = item_res.scalar_one_or_none()
  if item is None:
    raise HTTPException(status_code=404, detail='Строка не найдена')

  category = await _get_or_create_category(db, current_user.company_id, payload.category)
  brand = await _get_or_create_brand(db, current_user.company_id, payload.brand)
  product = await _get_or_create_product(
    db,
    brand.id,
    _norm_optional(payload.article),
    payload.name.strip(),
    _norm_optional(payload.size),
  )

  item.category_id = category.id
  item.product_id = product.id
  if payload.sortOrder is not None:
    item.sort_order = int(payload.sortOrder)

  await db.commit()

  out_res = await db.execute(
    select(ChecklistItem, RefCategory, RefProduct, RefBrand)
    .join(RefCategory, RefCategory.id == ChecklistItem.category_id)
    .join(RefProduct, RefProduct.id == ChecklistItem.product_id)
    .join(RefBrand, RefBrand.id == RefProduct.brand_id)
    .where(ChecklistItem.id == item.id),
  )
  out_item, out_category, out_product, out_brand = out_res.one()
  return ChecklistDetailsItem(
    itemId=out_item.id,
    sortOrder=out_item.sort_order,
    category=out_category.name,
    brand=out_brand.name,
    article=out_product.article,
    name=out_product.name,
    size=out_product.size,
  )


@router.delete(
  '/{project_id}/checklists/{checklist_id}/items/{item_id}',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Удалить строку чек-листа',
)
async def delete_checklist_item (
  project_id: UUID = Path(..., description='ID проекта'),
  checklist_id: UUID = Path(..., description='ID чек-листа'),
  item_id: UUID = Path(..., description='ID строки'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  await _get_project_or_404(db, current_user, project_id)

  cl_res = await db.execute(select(Checklist.id).where(Checklist.id == checklist_id, Checklist.project_id == project_id))
  if cl_res.scalar_one_or_none() is None:
    raise HTTPException(status_code=404, detail='Чек-лист не найден')

  item_res = await db.execute(select(ChecklistItem).where(ChecklistItem.id == item_id, ChecklistItem.checklist_id == checklist_id))
  item = item_res.scalar_one_or_none()
  if item is None:
    raise HTTPException(status_code=404, detail='Строка не найдена')

  await db.delete(item)
  await db.commit()
  return None


@router.patch(
  '/{project_id}/checklists/{checklist_id}',
  response_model=ChecklistDetailsResponse,
  summary='Обновить чек-лист',
)
async def update_checklist (
  payload: ChecklistUpdateRequest,
  project_id: UUID = Path(..., description='ID проекта'),
  checklist_id: UUID = Path(..., description='ID чек-листа'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> ChecklistDetailsResponse:
  await _get_project_or_404(db, current_user, project_id)

  res = await db.execute(
    select(Checklist)
    .options(selectinload(Checklist.chain))
    .where(Checklist.id == checklist_id, Checklist.project_id == project_id),
  )
  checklist = res.scalar_one_or_none()
  if checklist is None:
    raise HTTPException(status_code=404, detail='Чек-лист не найден')

  checklist.title = payload.title.strip()
  await db.commit()

  # вернуть обновлённый чек-лист с позициями
  return await get_checklist(project_id=project_id, checklist_id=checklist_id, db=db, current_user=current_user)


@router.delete(
  '/{project_id}/checklists/{checklist_id}',
  status_code=status.HTTP_204_NO_CONTENT,
  summary='Удалить чек-лист',
)
async def delete_checklist (
  project_id: UUID = Path(..., description='ID проекта'),
  checklist_id: UUID = Path(..., description='ID чек-листа'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> None:
  await _get_project_or_404(db, current_user, project_id)

  res = await db.execute(select(Checklist).where(Checklist.id == checklist_id, Checklist.project_id == project_id))
  checklist = res.scalar_one_or_none()
  if checklist is None:
    raise HTTPException(status_code=404, detail='Чек-лист не найден')

  await db.delete(checklist)
  await db.commit()
  return None


@router.get(
  '/{project_id}/points/{point_id}/catalog',
  response_model=PointCatalogResponse,
  summary='Каталог по точке (для анкеты)',
  description='Возвращает для выбранной точки сети/магазина бренды и товары из чек-листа.',
)
async def get_point_catalog (
  project_id: UUID = Path(..., description='ID проекта'),
  point_id: UUID = Path(..., description='ID точки'),
  db: AsyncSession = Depends(get_db),
  current_user: User = Depends(get_current_user),
) -> PointCatalogResponse:
  await _get_project_or_404(db, current_user, project_id)

  point_res = await db.execute(
    select(ShopPoint)
    .options(selectinload(ShopPoint.chain))
    .where(ShopPoint.id == point_id, ShopPoint.project_id == project_id),
  )
  point = point_res.scalar_one_or_none()
  if point is None:
    raise HTTPException(status_code=404, detail='Точка не найдена')

  checklist_res = await db.execute(
    select(Checklist)
    .options(selectinload(Checklist.chain))
    .where(Checklist.project_id == project_id, Checklist.chain_id == point.chain_id)
    .order_by(Checklist.created_at.desc()),
  )
  checklist = checklist_res.scalars().first()
  if checklist is None:
    return PointCatalogResponse(pointId=point.id, chain=to_chain_item(point.chain), brands=[])

  items_res = await db.execute(
    select(ChecklistItem, RefCategory, RefProduct, RefBrand)
    .join(RefCategory, RefCategory.id == ChecklistItem.category_id)
    .join(RefProduct, RefProduct.id == ChecklistItem.product_id)
    .join(RefBrand, RefBrand.id == RefProduct.brand_id)
    .where(ChecklistItem.checklist_id == checklist.id)
    .order_by(ChecklistItem.sort_order.asc()),
  )

  by_brand: dict[UUID, tuple[RefBrand, list[ChecklistProductItem]]] = {}
  for item, category, product, brand in items_res.all():
    bucket = by_brand.get(brand.id)
    if bucket is None:
      bucket = (brand, [])
      by_brand[brand.id] = bucket

    bucket[1].append(
      ChecklistProductItem(
        id=product.id,
        article=product.article,
        name=product.name,
        size=product.size,
        category=category.name,
      ),
    )

  brands_out: list[ChecklistBrandProducts] = []
  for _, data in by_brand.items():
    brand, products = data
    brands_out.append(
      ChecklistBrandProducts(
        brand=ChecklistBrandItem(id=brand.id, name=brand.name),
        products=products,
      ),
    )

  return PointCatalogResponse(pointId=point.id, chain=to_chain_item(point.chain), brands=brands_out)


@router.get(
  '/refs/regions',
  response_model=RefRegionsResponse,
  summary='Справочник регионов',
)
async def list_regions (
  db: AsyncSession = Depends(get_db),
) -> RefRegionsResponse:
  res = await db.execute(select(RefRegion).order_by(RefRegion.code.asc()))
  items = res.scalars().all()
  return RefRegionsResponse(items=[RefRegionItem(code=r.code, name=r.name) for r in items])


@router.get(
  '/refs/cities',
  response_model=RefCitiesResponse,
  summary='Справочник городов',
)
async def list_cities (
  regionCode: str | None = Query(default=None, description='Фильтр по коду региона'),
  db: AsyncSession = Depends(get_db),
) -> RefCitiesResponse:
  stmt = select(RefCity)
  if regionCode:
    code = regionCode.strip()
    if code.isdigit():
      code = _REGION_CODE_ALIASES.get(code, code)
      if len(code) == 3:
        candidate = code[-2:]
        exists = await db.execute(select(RefRegion.code).where(RefRegion.code == candidate))
        if exists.scalar_one_or_none() is not None:
          code = candidate
      if len(code) == 1:
        candidate = code.zfill(2)
        exists = await db.execute(select(RefRegion.code).where(RefRegion.code == candidate))
        if exists.scalar_one_or_none() is not None:
          code = candidate

    stmt = stmt.where(RefCity.region_code == code)
  stmt = stmt.order_by(RefCity.name.asc())

  res = await db.execute(stmt)
  items = res.scalars().all()
  return RefCitiesResponse(items=[RefCityItem(id=c.id, name=c.name, regionCode=c.region_code) for c in items])

